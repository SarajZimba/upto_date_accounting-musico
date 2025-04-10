from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.views.generic import CreateView,DetailView,ListView,TemplateView,UpdateView,View
from root.utils import DeleteMixin
from .models import AccountChart, Depreciation, FiscalYearLedger, FiscalYearSubLedger, CumulativeLedger
from django.views.generic import TemplateView
from .forms import AccountChartForm
from decimal import Decimal as D
from django.db.models import Q, Sum
from django.contrib import messages
from organization.models import Organization
from rest_framework.response import Response
from accounting.utils import calculate_depreciation
from rest_framework.decorators import api_view
from .utils import ProfitAndLossData
from user.permission import IsAdminMixin
import json
from decimal import Decimal
from itertools import chain
from operator import attrgetter
from datetime import datetime
from django.db import models
from .forms import JournalEntryForm
from collections import defaultdict
from django.db.models.functions import Coalesce
from accounting.utils import change_date_to_datetime




class AccountChartMixin(IsAdminMixin):
    model = AccountChart
    form_class = AccountChartForm
    paginate_by = 10
    queryset = AccountChart.objects.prefetch_related('accountledger_set')
    success_url = reverse_lazy('accountchart_list')


class AccountChartList(AccountChartMixin, ListView):
    queryset = AccountChart.objects.all()
    template_name = "accounting/accounting_chart.html"


    def get(self, request, *args, **kwargs):
        query_set = self.queryset
        sundry_debtors_total = AccountLedger.objects.filter(account_chart__group='Sundry Debtors').aggregate(total_value_sum=Sum('total_value'))['total_value_sum']
        sundry_creditors_total = AccountLedger.objects.filter(account_chart__group='Sundry Creditors').aggregate(total_value_sum=Sum('total_value'))['total_value_sum']
        # print(sundry_debtors_total)
        assets = query_set.filter(account_type='Asset')
        liabilities = query_set.filter(account_type='Liability')
        equities = query_set.filter(account_type='Equity')
        revenues = query_set.filter(account_type='Revenue')
        expenses = query_set.filter(account_type='Expense')
        others = query_set.filter(account_type='Other')


        context = {
            'sundry_debtors_total': sundry_debtors_total,
            'sundry_creditors_total': sundry_creditors_total,
            'assets': assets,
            'liabilities':liabilities,
            'equities':equities,
            'revenues':revenues,
            'expenses': expenses,
            'others': others
        }
        return render(request, 'accounting/accounting_chart.html', context)



class AccountChartDetail(AccountChartMixin, DetailView):
    template_name = "accounting/accountchart_detail.html"

class AccountChartCreate(AccountChartMixin, CreateView):
    template_name = "accounting/create.html"

class AccountChartUpdate(AccountChartMixin, UpdateView):
    template_name = "update.html"

class AccountChartDelete(AccountChartMixin, DeleteMixin, View):
    pass


from .models import AccountLedger, AccountSubLedger
from .forms import AccountLedgerForm
class AccountLedgerMixin(IsAdminMixin):
    model = AccountLedger
    form_class = AccountLedgerForm
    paginate_by = 10
    queryset = AccountLedger.objects.all()
    success_url = reverse_lazy('accountledger_list')

class AccountLedgerList(AccountLedgerMixin, ListView):
    template_name = "accounting/accountledger_list.html"
    queryset = AccountLedger.objects.all()

class AccountLedgerDetail(AccountLedgerMixin, DetailView):
    template_name = "accounting/accountledger_detail.html"

from bill.utils import update_cumulative_ledger_bill, create_cumulative_ledger_bill
class AccountLedgerCreate(AccountLedgerMixin, CreateView):
    template_name = "accounting/create.html"

    def form_valid(self, form):
        response = super().form_valid(form)
        entry_date = datetime.now()
        create_cumulative_ledger_bill(self.object, entry_date)
        return response

class AccountLedgerUpdate(AccountLedgerMixin, UpdateView):
    template_name = "accounting/update.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        print(self.object)
        ledger = self.object
        ledger_id = self.object.id
        ledger_name = self.object.ledger_name
        opening_count =ledger.opening_count
        opening_exists = False
        context["opening_count"] = opening_count 

        not_first_entry = CumulativeLedger.objects.filter(ledger=self.object, total_value__gt = 0.0).exists()
        if not_first_entry:
            context['no_transaction_exists'] = False
        else:
            context['no_transaction_exists'] = True  


        return context
    
    def form_valid(self, form):
        response = super().form_valid(form)


        entry_datetime_forcumulativeledger = datetime.now()
        update_cumulative_ledger_bill(self.object, entry_datetime_forcumulativeledger)
        not_first_entry = CumulativeLedger.objects.filter(ledger=self.object).exists()
        if not_first_entry and self.object.opening_count == Decimal(0.0):
            self.object.opening_count = self.object.total_value
            self.object.save()
        return response

class AccountLedgerDelete(View):
    model = AccountLedger
    def get(self, request, pk):
        object = self.model.objects.get(id=pk)
        object.delete()
        return redirect(reverse_lazy('accountledger_list'))

from django.db import transaction

from .forms import AccountSubLedgerForm
class AccountSubLedgerCreate(IsAdminMixin, CreateView):
    template_name = "accounting/subledger/create.html"
    form_class = AccountSubLedgerForm
    success_url = reverse_lazy('accountchart_list')

    def form_valid(self, form):
        # Use a transaction to ensure atomicity
        with transaction.atomic():
            # Get the form data
            response = super().form_valid(form)
            
            # Get the newly created subledger object
            subledger = self.object
            
            # Since it's a new creation, prev_value is 0
            prev_value = 0
            new_value = subledger.total_value
            
            # Create a tracking entry for the new subledger
            AccountSubLedgerTracking.objects.create(
                subledger=subledger,
                prev_amount=prev_value,
                new_amount=new_value,
                value_changed=new_value - prev_value
            )

            ledger = subledger.ledger

            ledger.total_value = ledger.total_value + new_value
            ledger.save()
            entry_date = datetime.now()
            update_cumulative_ledger_bill(ledger, entry_date)
            
            return response

class AccountSubLedgerUpdate(IsAdminMixin, UpdateView):
    form_class = AccountSubLedgerForm
    queryset = AccountSubLedger.objects.all()
    template_name = "update.html"
    success_url = reverse_lazy('accountchart_list')

    def form_valid(self, form):
        # Use a transaction to ensure atomicity
        with transaction.atomic():
            pk = self.kwargs['pk']
            
            # Fetch the original object from the database
            subledger = get_object_or_404(AccountSubLedger, pk=pk)
            
            # Capture the previous total value from the original object
            prev_value = subledger.total_value
            # Fetch the previous total value from the object before the update
            prev_value = subledger.total_value
            print(prev_value)
            # Apply the form data to the object
            response = super().form_valid(form)
            
            subledger.refresh_from_db()
            # The object has been updated with the form data
            # Fetch the new total value after saving
            new_value = subledger.total_value
            


            value_changed=new_value - prev_value

            print(value_changed)
            # Create a tracking entry for the subledger
            subledgertracking = AccountSubLedgerTracking.objects.create(
                subledger=subledger,
                prev_amount=prev_value,
                new_amount=new_value,
                value_changed= value_changed
            )
            ledger = subledger.ledger
            ledger_prev_value = ledger.total_value
            if value_changed:
                # ledger_prev_value
                ledger.total_value = ledger_prev_value + value_changed
                ledger.save()
                entry_date = datetime.now()
                update_cumulative_ledger_bill(ledger, entry_date)
            return response

from .models import Expense
from .forms import ExpenseForm
class ExpenseMixin(IsAdminMixin):
    model = Expense
    form_class = ExpenseForm
    paginate_by = 10
    queryset = Expense.objects.all().order_by('-id')
    success_url = reverse_lazy('expenses_list')

class ExpenseList(ExpenseMixin, ListView):
    template_name = "accounting/expenses/expenses_list.html"

class ExpenseDetail(ExpenseMixin, DetailView):
    template_name = "expense/expense_detail.html"

# class ExpenseCreate(ExpenseMixin, CreateView):
#     template_name = "accounting/expenses/expenses_create.html"

from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views import View
from .models import Expense, AccountLedger, AccountSubLedger
from django.http import JsonResponse
from rest_framework.views import APIView
from rest_framework.response import Response
# from .serializers import AccountSubLedgerSerializer
from accounting.models import AccountSubLedgerTracking
from bill.utils import update_cumulative_ledger_expense


import decimal
class ExpenseCreate(View):
    template_name = "accounting/expenses/expenses_create.html"
    success_url = reverse_lazy('expenses_list')
    def get(self, request, *args, **kwargs):
        context = {
            # 'ledgers': AccountLedger.objects.filter(account_chart__account_type="Expense").exclude(ledger_name='Inventory Expenses'),
            'ledgers': AccountLedger.objects.filter(account_chart__account_type="Expense"),
            'credit_ledgers': AccountLedger.objects.filter(account_chart__group="Liquid Asset"),
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        print(request.POST)
        ledger_id = request.POST.get('ledger')
        sub_ledger_id = request.POST.get('sub_ledger')
        credit_ledger_id = request.POST.get('credit_ledger')
        credit_sub_ledger_id = request.POST.get('credit_sub_ledger')
        amount = int(request.POST.get('amount'))
        description = request.POST.get('description')
        entry_date = request.POST.get('entry_date')

        entry_datetime_for_cumulativeledger = change_date_to_datetime(entry_date)
        

        print(f"after function change {entry_datetime_for_cumulativeledger}")
        if ledger_id != '':
            ledger = AccountLedger.objects.get(id=ledger_id)
        else:
            ledger= None
        if sub_ledger_id != '' :
            sub_ledger = AccountSubLedger.objects.get(id=sub_ledger_id, ledger = ledger)

        else:
            sub_ledger = None
        if credit_ledger_id != '':
            credit_ledger = AccountLedger.objects.get(id=credit_ledger_id)
        else:
            credit_ledger = None
        if credit_sub_ledger_id != '' :
            credit_sub_ledger = AccountSubLedger.objects.get(id=credit_sub_ledger_id, ledger = credit_ledger)

        else:
            credit_sub_ledger = None

        journal = TblJournalEntry.objects.create(employee_name="From expense form", journal_total=amount, entry_date=entry_date)
        TblDrJournalEntry.objects.create(ledger=ledger, debit_amount=amount, particulars=f"Automatic: {ledger.ledger_name} a/c Dr", journal_entry=journal, sub_ledger=sub_ledger, paidfrom_ledger = credit_ledger)
        TblCrJournalEntry.objects.create(ledger=credit_ledger, credit_amount=amount, particulars=f"Automatic: To {credit_ledger.ledger_name}", journal_entry=journal, sub_ledger=credit_sub_ledger,paidfrom_ledger = credit_ledger)
        instance = Expense.objects.create(
            ledger=ledger,
            sub_ledger=sub_ledger,
            credit_ledger=credit_ledger,
            credit_sub_ledger=credit_sub_ledger,
            amount=amount,
            description=description, 
            journal=journal,
            entry_date=entry_date
        )


        instance.ledger.total_value += instance.amount
        instance.ledger.save()
        update_cumulative_ledger_expense(instance.ledger, journal, entry_datetime_for_cumulativeledger)

        instance.credit_ledger.total_value -= instance.amount
        instance.credit_ledger.save()
        update_cumulative_ledger_expense(instance.credit_ledger, journal, entry_datetime_for_cumulativeledger)

        expense_subledger = instance.sub_ledger
        
        if expense_subledger:
        
            prev_value = expense_subledger.total_value
            subledgertracking = AccountSubLedgerTracking.objects.create(subledger = sub_ledger, prev_amount= prev_value, journal=journal)
            expense_subledger.total_value += decimal.Decimal(int(amount))
            expense_subledger.save()
            subledgertracking.new_amount=expense_subledger.total_value
            subledgertracking.value_changed = expense_subledger.total_value - prev_value
            subledgertracking.save()

        expense_credit_subledger = instance.credit_sub_ledger
            # sub = AccountSubLedger.objects.get(sub_ledger_name=subledgername, ledger=debit_account)
        if expense_credit_subledger:
            prev_value = expense_credit_subledger.total_value
            credit_subledgertracking = AccountSubLedgerTracking.objects.create(subledger = credit_sub_ledger, prev_amount= prev_value, journal=journal)
            expense_credit_subledger.total_value -= decimal.Decimal(int(amount))
            expense_credit_subledger.save()
            credit_subledgertracking.new_amount=expense_credit_subledger.total_value
            credit_subledgertracking.value_changed = expense_credit_subledger.total_value - prev_value
            credit_subledgertracking.save()
        return redirect(reverse_lazy('expenses_list'))



class ExpenseUpdate(ExpenseMixin, UpdateView):
    template_name = "update.html"


from .utils import soft_delete_journal_expense
class ExpenseDelete(View):
    model = Expense
    # success_url = reverse_lazy('expenses_list')
    def get(self, request, pk):
        # obj = get_object_or_404(self.model, id=pk)
        object = self.model.objects.get(id=pk)
        journal_entry=object.journal
        # adjust_cumulative_ledger_afterentries(journal_entry)
        soft_delete_journal_expense(journal_entry)
        object.delete()
        # object.save()
        return redirect(reverse_lazy('expenses_list'))


from .models import TblDrJournalEntry, TblCrJournalEntry, TblJournalEntry, AccountSubLedger
from .utils import update_cumulative_ledger_journal, create_cumulative_ledger_journal, change_date_to_datetime

class JournalEntryCreateView(IsAdminMixin,View):

    # def get(self, request):
    #     ledgers = AccountLedger.objects.all()
    #     sub_ledgers = AccountSubLedger.objects.all()
    #     return render(request, 'accounting/journal/journal_entry_create.html', {'ledgers':ledgers, 'sub_ledgers':sub_ledgers})
    
    def get(self, request):
        ledgers = AccountLedger.objects.all()
        sub_ledgers = AccountSubLedger.objects.all()

        ledgers_list = []
        for ledger in ledgers:
            ledgers_dict = {
                "id" : ledger.id,
                "account_chart" : ledger.account_chart.id,
                "ledger_name" : ledger.ledger_name,
                "total_value" : float(ledger.total_value),
                "is_editable" : ledger.is_editable,
                "has_subledgers": AccountSubLedger.objects.filter(ledger=ledger).exists()
            }

            ledgers_list.append(ledgers_dict)

        # print(ledgers_list)
        return render(request, 'accounting/journal/journal_entry_create.html', {'ledgers':ledgers_list, 'sub_ledgers':sub_ledgers,'ledgers_list_json': json.dumps(ledgers_list)})
    

    def get_subledger(self, subledger, ledger):
        subled = None
        if not subledger.startswith('-'):
            try:
                subledger_id = int(subledger)
                subled = AccountSubLedger.objects.get(pk=subledger_id)
            except ValueError:
                subled = AccountSubLedger.objects.create(sub_ledger_name=subledger, is_editable=True, ledger=ledger)
        return subled

    def post(self, request):
        data = request.POST
        debit_ledgers = data.getlist('debit_ledger', [])
        debit_particulars = data.getlist('debit_particular', [])
        debit_amounts = data.getlist('debit_amount', [])
        debit_subledgers = data.getlist('debit_subledger', [])

        credit_ledgers = data.getlist('credit_ledger', [])
        credit_particulars = data.getlist('credit_particular', [])
        credit_amounts = data.getlist('credit_amount', [])
        credit_subledgers = data.getlist('credit_subledger', [])
        # print(credit_ledgers)
        entry_date = data.get('entry_date')
        entry_datetime_for_cumulativeledger = change_date_to_datetime(entry_date)
        narration = data.get('narration')
        ledgers = AccountLedger.objects.all()
        sub_ledgers = AccountSubLedger.objects.all()

        try:
            parsed_debitamt = (lambda x: [D(i) for i in x])(debit_amounts)
            parsed_creditamt = (lambda x: [D(i) for i in x])(credit_amounts)
        except Exception:
            messages.error(request, "Please Enter a valid amount")
            return render(request, 'accounting/journal/journal_entry_create.html', {'ledgers': ledgers, 'sub_ledgers': sub_ledgers})

        debit_sum, credit_sum = sum(parsed_debitamt), sum(parsed_creditamt)
        if debit_sum != credit_sum:
            messages.error(request, "Debit Total and Credit Total must be equal")
            return render(request, 'accounting/journal/journal_entry_create.html', {'ledgers': ledgers, 'sub_ledgers': sub_ledgers})

        for dr in debit_ledgers:
            if dr.startswith('-'):
                messages.error(request, "Ledger must be selected")
                return render(request, 'accounting/journal/journal_entry_create.html', {'ledgers': ledgers, 'sub_ledgers': sub_ledgers})

        credit_to_debit_mapping = {}

        journal_entry = TblJournalEntry.objects.create(employee_name=request.user.username, journal_total=debit_sum,  entry_date=entry_date, narration=narration)
        for i in range(len(credit_ledgers)):
            credit_ledger_id = int(credit_ledgers[i])
            
            credit_ledger = AccountLedger.objects.get(pk=credit_ledger_id)
          
            credit_to_debit_mapping[credit_ledger] = credit_ledger
            credit_particular = credit_particulars[i]
            credit_amount = D(credit_amounts[i])
            subledger = self.get_subledger(credit_subledgers[i], credit_ledger)
            credit_ledger_type = credit_ledger.account_chart.account_type
            TblCrJournalEntry.objects.create(ledger=credit_ledger, journal_entry=journal_entry, particulars=credit_particular, credit_amount=credit_amount, sub_ledger=subledger, paidfrom_ledger=credit_ledger)
            if credit_ledger_type in ['Asset', 'Expense']:
                credit_ledger.total_value = credit_ledger.total_value - credit_amount
                credit_ledger.save()
                create_cumulative_ledger_journal(credit_ledger, journal_entry, entry_datetime_for_cumulativeledger)
                # create_cumulative_ledger_journal(credit_ledger)
                if subledger:
                    prev_value = subledger.total_value
                    subledgertracking = AccountSubLedgerTracking.objects.create(subledger = subledger, prev_amount= prev_value, journal=journal_entry )

                    subledger.total_value = subledger.total_value - credit_amount
                    subledger.save()
                    subledgertracking.new_amount=subledger.total_value
                    subledgertracking.value_changed = subledger.total_value - prev_value
                    subledgertracking.save()
            elif credit_ledger_type in ['Liability', 'Revenue', 'Equity']:
                credit_ledger.total_value = credit_ledger.total_value + credit_amount
                credit_ledger.save()
                create_cumulative_ledger_journal(credit_ledger, journal_entry, entry_datetime_for_cumulativeledger)
                # create_cumulative_ledger_journal(credit_ledger)
                if subledger:
                    prev_value = subledger.total_value
                    subledgertracking = AccountSubLedgerTracking.objects.create(subledger = subledger, prev_amount= prev_value, journal=journal_entry)

                    subledger.total_value = subledger.total_value + credit_amount
                    subledger.save()
                    subledgertracking.new_amount=subledger.total_value
                    subledgertracking.value_changed = subledger.total_value - prev_value
                    subledgertracking.save()

        
        for i in range(len(debit_ledgers)):
            debit_ledger_id = int(debit_ledgers[i])
            debit_ledger = AccountLedger.objects.get(pk=debit_ledger_id)
            debit_particular = debit_particulars[i]
            debit_amount = D(debit_amounts[i])
            subledger = self.get_subledger(debit_subledgers[i], debit_ledger)
            debit_ledger_type = debit_ledger.account_chart.account_type
            TblDrJournalEntry.objects.create(ledger=debit_ledger, journal_entry=journal_entry, particulars=debit_particular, debit_amount=debit_amount, sub_ledger=subledger, paidfrom_ledger=credit_to_debit_mapping.get(credit_ledger))
            if debit_ledger_type in ['Asset', 'Expense']:
                debit_ledger.total_value = debit_ledger.total_value + debit_amount
                debit_ledger.save()
                create_cumulative_ledger_journal(debit_ledger, journal_entry, entry_datetime_for_cumulativeledger)
                # create_cumulative_ledger_journal(debit_ledger)
                if subledger:
                    prev_value = subledger.total_value
                    subledgertracking = AccountSubLedgerTracking.objects.create(subledger = subledger, prev_amount= prev_value, journal=journal_entry)

                    subledger.total_value = subledger.total_value + debit_amount
                    subledger.save()
                    subledgertracking.new_amount=subledger.total_value
                    subledgertracking.value_changed = subledger.total_value - prev_value
                    subledgertracking.save()
            elif debit_ledger_type in ['Liability', 'Revenue', 'Equity']:
                debit_ledger.total_value = debit_ledger.total_value - debit_amount
                debit_ledger.save()
                create_cumulative_ledger_journal(debit_ledger, journal_entry, entry_datetime_for_cumulativeledger)
                if subledger:
                    prev_value = subledger.total_value
                    subledgertracking = AccountSubLedgerTracking.objects.create(subledger = subledger, prev_amount= prev_value, journal=journal_entry)

                    subledger.total_value = subledger.total_value - debit_amount
                    subledger.save()
                    subledgertracking.new_amount=subledger.total_value
                    subledgertracking.value_changed = subledger.total_value - prev_value
                    subledgertracking.save()

        return redirect('journal_list')



class JournalEntryView(IsAdminMixin, View):

    def get(self, request, pk=None):
        from_date = request.GET.get('fromDate', None)
        to_date = request.GET.get('toDate', None)

        if from_date and to_date and (to_date > from_date):
            journals = TblJournalEntry.objects.filter(created_at__range=[from_date, to_date])

            journal_entries = {
                'entries': [],
                "debit_sum": 0,
                "credit_sum": 0
            }
            debit_sum, credit_sum = 0,0
            for journal in journals:
                data = {'dr':[], 'cr':[], "dr_total": 0, "cr_total": 0}
                for dr in journal.tbldrjournalentry_set.all():
                    data['dr'].append(dr)
                    data['dr_total'] += dr.debit_amount
                for cr in journal.tblcrjournalentry_set.all():
                    data['cr'].append(cr)
                    data['cr_total'] += cr.credit_amount
                journal_entries['entries'].append(data)
                journal_entries['debit_sum']+=data['dr_total']
                journal_entries['credit_sum']+=data['cr_total']


            context = {
                'from_date':from_date,
                'to_date': to_date,
                'journals':journal_entries
            }

            return render(request,'accounting/journal/journal.html' , context=context)
        if pk:
            journal = TblJournalEntry.objects.get(pk=pk)
            credit_details = TblCrJournalEntry.objects.filter(journal_entry=journal)
            debit_details = TblDrJournalEntry.objects.filter(journal_entry=journal)
            debit_total, credit_total = 0, 0
            for dr in debit_details:
                debit_total += dr.debit_amount

            for cr in credit_details:
                credit_total += cr.credit_amount
            check_cumulative = CumulativeLedger.objects.filter(journal=journal).exists()
            print("This is the value of check_cumulative", check_cumulative)
            context = {
                'credit': credit_details,
                'debit': debit_details,
                "dr_total":debit_total,
                "cr_total": credit_total,
                'journal':journal,
                'is_updatable':check_cumulative
            }
            return render(request, 'accounting/journal/journal_voucher.html', context)
            

        journal_entries = TblJournalEntry.objects.prefetch_related('tbldrjournalentry_set').order_by('-created_at').all()
        return render(request, 'accounting/journal/journal_list.html',  {'journal_entries': journal_entries})



from .detail_utils import give_detail, get_standard_trial_balance
class TrialBalanceView(IsAdminMixin, View):

    def filtered_view(self, from_date, to_date):
        filtered_transactions = CumulativeLedger.objects.filter(entry_date__range=[from_date, to_date])
        filtered_sum = filtered_transactions.values('ledger_name', 'account_chart__account_type', 'account_chart__group').annotate(Sum('value_changed'))
        # print(filtered_sum)
        trial_balance = []

        total = {'debit_total':0, 'credit_total':0}

        for fil in filtered_sum:
            data = {}
            data['ledger'] = fil['ledger_name']
            account_type = fil['account_chart__account_type']
            account_group = fil['account_chart__group']
            # print(account_group)
            if account_type in ['Asset', 'Expense']:
                data['actual_value'] = fil['value_changed__sum']
                if fil['value_changed__sum'] < 0:
                    val = abs(fil['value_changed__sum'])
                    data['credit'] = val
                    data['debit'] = '-'
                    total['credit_total'] += val
                else:
                    val = fil['value_changed__sum']
                    data['debit'] = val
                    data['credit'] = '-'
                    total['debit_total'] += val
            else:
                if fil['value_changed__sum'] < 0:
                    val = abs(fil['value_changed__sum'])
                    data['debit'] = val
                    data['credit'] = '-'
                    total['debit_total'] += val
                else:
                    val = fil['value_changed__sum']
                    data['credit'] = val
                    data['debit'] = '-'
                    total['credit_total'] += val

            if not any(d['account_type'] == account_type for d in trial_balance):
                    trial_balance.append(
                        {
                            'account_type': account_type,
                            'ledgers' : [data],
                            'group' : account_group,
                        }
                    )
            else:
                for tb in trial_balance:
                    if tb['account_type'] == account_type:
                        # print(data)
                        tb['ledgers'].append(data)
                        break

        return trial_balance, total

    def detail_view(self, from_date, to_date):
        all_ledgers_list = AccountLedger.objects.values_list('ledger_name', flat=True)
        before_transactions = CumulativeLedger.objects.filter(entry_date__lt=from_date, total_value__gt=0).order_by('-created_at')

        trial_balance = []
        total = {'debit_total':0, 'credit_total':0}

        filtered_transactions = CumulativeLedger.objects.filter(entry_date__range=[from_date, to_date])
        filtered_sum = filtered_transactions.values('ledger_name', 'account_chart__account_type', 'account_chart__group' ).annotate(Sum('debit_amount'), Sum('credit_amount'), Sum('value_changed'))

        for fil in filtered_sum:
            data = {}
            data['ledger'] = fil['ledger_name']
            account_type = fil['account_chart__account_type']
            account_group = fil['account_chart__account_group']
            data['debit'] = fil['debit_amount__sum']
            data['credit'] = fil['credit_amount__sum']
            if account_type in ['Asset', 'Expense']:
                if fil['value_changed__sum'] < 0:
                    total['credit_total'] += abs(fil['value_changed__sum'])
                else:
                    total['debit_total'] += abs(fil['value_changed__sum'])
            else:
                if fil['value_changed__sum'] < 0:
                    total['debit_total'] += abs(fil['value_changed__sum'])
                else:
                    total['credit_total'] += abs(fil['value_changed__sum'])



            if not any(d['account_type'] == account_type for d in trial_balance):
                    trial_balance.append(
                        {
                            'account_type': account_type,
                            'ledgers' : [data],
                            'group' : account_group
                        }
                    )
            else:
                for tb in trial_balance:
                    if tb['account_type'] == account_type:
                        tb['ledgers'].append(data)
                        break


        included_ledgers = []

        for trans in before_transactions:
            account_type = trans.account_chart.account_type
            if trans.ledger_name not in included_ledgers:
                included_ledgers.append(trans.ledger_name)
                if not any(d['account_type'] == account_type for d in trial_balance):
                    data = {
                        'ledger': trans.ledger_name,
                        'opening': trans.total_value,
                        'debit':'-',
                        'credit':'-',
                        'closing': trans.total_value,
                        'group': account_group
                    }
                    trial_balance.append({'account_type':account_type, 'ledgers':[data], 'group': account_group})
                else:
                    for tb in trial_balance:
                        if tb['account_type'] == account_type:
                            if not any(d['ledger'] == trans.ledger_name for d in tb['ledgers']):
                                tb['ledgers'].append({
                                    'ledger': trans.ledger_name,
                                    'opening': trans.total_value,
                                    'debit':'-',
                                    'credit':'-',
                                    'closing': trans.total_value,
                                    'group': account_group
                                })
                            else:
                                for led in tb['ledgers']:
                                    if led['ledger'] == trans.ledger_name:
                                        led['opening'] = trans.total_value
                                        if account_type in ['Asset', 'Expense']:
                                            led['closing'] = trans.total_value + led['debit'] - led['credit']
                                        else:
                                            led['closing'] = trans.total_value + led['credit'] - led['debit']
                                        break


            if len(included_ledgers) >= len(all_ledgers_list):
                break


 
        return trial_balance, total


    def get(self, request):
        from_date = request.GET.get('fromDate', None)
        to_date = request.GET.get('toDate', None)
        option = request.GET.get('option', None)
        current_fiscal_year = Organization.objects.last().current_fiscal_year
        first_date=None
        last_date=None

        if from_date and to_date:
            if option and option =='openclose':
                # trial_balance, total = self.detail_view(from_date, to_date)
                trial_balance, total = give_detail(from_date, to_date)
                context = {
                    'trial_balance': trial_balance,
                    "total": total,
                    "from_date":from_date,
                    "to_date":to_date,
                    'openclose':True,
                    'current_fiscal_year':current_fiscal_year,
                    'trial_active' : Organization.objects.first().show_zero_ledgers
                }
                return render(request, 'accounting/trial_balance.html', context)
            else:
                trial_balance, total= self.filtered_view(from_date, to_date)
                context = {
                    'trial_balance': trial_balance,
                    "total": total,
                    "from_date":from_date,
                    "to_date":to_date,
                    'current_fiscal_year':current_fiscal_year,
                    'trial_active' : Organization.objects.first().show_zero_ledgers

                }

                return render(request, 'accounting/trial_balance.html', context)
        
        else:

                trial_balance = []
                total = {'debit_total': 0, 'credit_total': 0}
                ledgers = AccountLedger.objects.filter(~Q(total_value=0), ~Q(ledger_name='VAT Receivable'), ~Q(ledger_name='VAT Payable'))
                # ledgers = AccountLedger.objects.filter(~Q(total_value=0))

                # Create a dictionary to store ledger entries based on their names
                ledger_dict = {}

                for led in ledgers:
                    data = {}
                    account_type = led.account_chart.account_type
                    account_group = led.account_chart.group
                    data['account_type'] = account_type
                    data['ledger'] = led.ledger_name
                    data['group'] = account_group  # Use account_group here
                    data['id'] = led.id

                    if Organization.objects.first().show_zero_ledgers:
                        subledgers = AccountSubLedger.objects.filter(ledger_id = led.id)
                    else:
                        subledgers = AccountSubLedger.objects.filter(ledger_id = led.id, total_value__gt=0.0)
                    # subledgers = AccountSubLedger.objects.filter(ledger_id = led.id)
    
                    subledger_entries = []

                    for subledger in subledgers:
                        subledger_data = {
                            'subledger': subledger.sub_ledger_name,
                            'total_value': subledger.total_value
                        }
                        subledger_entries.append(subledger_data)
                    data['subledgers'] = subledger_entries
                    if account_type in ['Asset', 'Expense']:
                        if led.total_value > 0:
                            data['debit'] = led.total_value
                            if account_group not in ["Sundry Debtors", "Sundry Creditors"]:
                                total['debit_total'] += led.total_value
                            data['credit'] = '-'
                        else:
                            val = abs(led.total_value)
                            data['credit'] = val
                            if account_group not in ["Sundry Debtors", "Sundry Creditors"]:
                                total['credit_total'] += val
                            data['debit'] = '-'
                    else:
                        if led.total_value > 0:
                            data['credit'] = led.total_value
                            if account_group not in ["Sundry Debtors", "Sundry Creditors"]:
                                total['credit_total'] += led.total_value
                            data['debit'] = '-'
                        else:
                            val = abs(led.total_value)
                            data['debit'] = val
                            if account_group not in ["Sundry Debtors", "Sundry Creditors"]:
                                total['debit_total'] += val
                            data['credit'] = '-'

                    # Use the ledger name as a key to avoid redundancy
                    ledger_name = data['ledger']
                    if ledger_name not in ledger_dict:
                        ledger_dict[ledger_name] = data

                # Convert the dictionary values to a list
                ledger_entries = list(ledger_dict.values())

                for entry in ledger_entries:
                    real_account_type = entry['account_type']
                    account_type = entry['group']
                    account_group = entry['group']


                    # Check if the account_type already exists in trial_balance
                    account_type_entry = next((d for d in trial_balance if d['account_type'] == account_type), None)

                    if not account_type_entry:
                        trial_balance.append({'real_account_type': real_account_type, 'account_type': account_type, 'groups': [{'group': account_group, 'ledgers': [entry]}]})
                    else:
                        # Find the dictionary for the current account_type
                        current_account_type = next((d for d in trial_balance if d['account_type'] == account_type), None)

                        # Check if the account_group already exists for the current account_type
                        account_group_entry = next((g for g in current_account_type['groups'] if g['group'] == account_group), None)

                        if not account_group_entry:
                            current_account_type['groups'].append({'group': account_group, 'ledgers': [entry]})
                        else:
                            account_group_entry['ledgers'].append(entry)

                trial_balance = [dict(item) for item in trial_balance]
                vat_receivable = AccountLedger.objects.get(ledger_name= 'VAT Receivable')
                vat_payable = AccountLedger.objects.get(ledger_name= 'VAT Payable')



                vat_receivable_val = vat_receivable.total_value
                vat_payable_val = vat_payable.total_value

                print(f"vat_receivable {vat_receivable_val}")
                print(f"vat_payable {vat_payable_val}")

                vat = vat_receivable_val - vat_payable_val
                if vat > 0:
                    vatreceivable_subledgers = AccountSubLedger.objects.filter(ledger__ledger_name= 'VAT Receivable')
                    vatreceivable_subledger_entries = []
                    for subledger in vatreceivable_subledgers:
                        subledger_data = {
                                'subledger': subledger.sub_ledger_name,
                                'total_value': subledger.total_value
                            }
                        vatreceivable_subledger_entries.append(subledger_data)  
                    new_vat_receivable_entry = {'real_account_type': vat_receivable.account_chart.account_type, 'account_type': vat_receivable.account_chart.group, 'groups': [{'group': vat_receivable.account_chart.group,"ledgers": [{'ledger': vat_receivable.ledger_name, 'group': vat_receivable.account_chart.account_type, 'debit': abs(vat), 'credit': 0, 'subledgers': vatreceivable_subledger_entries}] }]}
                    total['debit_total'] += abs(vat)
                    # total['credit_total'] += abs(vat)
                    trial_balance.append(new_vat_receivable_entry)
                else:
                    vatpayable_subledgers = AccountSubLedger.objects.filter(ledger__ledger_name= 'VAT Payable')
                    vatpayable_subledger_entries = []
                    for subledger in vatpayable_subledgers:
                        subledger_data = {
                                'subledger': subledger.sub_ledger_name,
                                'total_value': subledger.total_value
                            }
                        vatpayable_subledger_entries.append(subledger_data)  
                    new_vat_payable_entry = {'real_account_type': vat_payable.account_chart.account_type, 'account_type': vat_payable.account_chart.group, 'groups': [{'group': vat_payable.account_chart.group,"ledgers": [{'ledger': vat_payable.ledger_name, 'group': vat_payable.account_chart.account_type, 'debit': 0, 'credit': abs(vat), 'subledgers': vatpayable_subledger_entries}] }]}
                    total['credit_total'] += abs(vat)
                    trial_balance.append(new_vat_payable_entry)


                trial_balance.sort(key=lambda x: x['real_account_type'])

        Sundry_debtors_debit_total = 0
        Sundry_debtors_credit_total = 0
        Sundry_creditors_debit_total = 0
        Sundry_creditors_credit_total = 0
        new_ledgers_asset = []
        new_ledgers_liability = []


        for entry in trial_balance:
            new_ledger_entries = []
            for datas in entry['groups']:
                for data in datas['ledgers']:
                    group = data.get('group')
                    if group == 'Sundry Debtors':
                        if(data.get('debit') != '-'):
                            Sundry_debtors_debit_total += data.get('debit')

                        if(data.get('credit') != '-'):    
                            Sundry_debtors_credit_total += data.get('credit')

                    elif group == 'Sundry Creditors':
                        if(data.get('credit') != '-'):
                                Sundry_creditors_credit_total += data.get('credit')
                        if(data.get('debit') != '-'):
                                Sundry_creditors_debit_total += data.get('debit')
                    else:
                        new_ledger_entries.append(data)  # Keep all other ledgers
                
            datas['ledgers'] = new_ledger_entries  # Replace ledgers with the filtered list

        # Create new ledger entries for Sundry Debtors and add to the Asset section
        if (Sundry_debtors_debit_total != 0) or (Sundry_debtors_credit_total !=0):
            Sundry_debtors_debit_total_after = 0
            Sundry_debtors_credit_total_after = 0

            sundry_debtors_total = Sundry_debtors_debit_total - Sundry_debtors_credit_total
            if sundry_debtors_total > 0:
                Sundry_debtors_debit_total_after = sundry_debtors_total
                total['debit_total'] += Sundry_debtors_debit_total_after
                Sundry_debtors_credit_total_after = "-"
            else:
                Sundry_debtors_debit_total_after = "-"
                Sundry_debtors_credit_total_after = abs(sundry_debtors_total)
                total['credit_total'] += Sundry_debtors_credit_total_after
                
            ledgers = AccountLedger.objects.filter(account_chart__group="Sundry Debtors")
            subledger_entries = []

            for ledger in ledgers:
                subledger_data = {
                    'subledger': ledger.ledger_name,
                    'total_value': ledger.total_value
                }
                subledger_entries.append(subledger_data)
                
            new_sundry_debtors_entry = {'ledger': 'Sundry Debtors', 'group': 'Asset', 'debit': Sundry_debtors_debit_total_after, 'credit': Sundry_debtors_credit_total_after, 'subledgers': subledger_entries}
            new_ledgers_asset.append(new_sundry_debtors_entry)

        # Create new ledger entries for Sundry Creditors and add to the Liability section
        if (Sundry_creditors_debit_total != 0) or (Sundry_creditors_credit_total!= 0):
            Sundry_creditors_debit_total_after = 0
            Sundry_creditors_credit_total_after = 0

            sundry_creditors_total = Sundry_creditors_credit_total - Sundry_creditors_debit_total
            if sundry_creditors_total > 0:
                Sundry_creditors_debit_total_after = "-" 
                Sundry_creditors_credit_total_after = sundry_creditors_total
                total['credit_total'] += Sundry_creditors_credit_total_after
            else:
                Sundry_creditors_debit_total_after = abs(sundry_creditors_total)
                total['debit_total'] += Sundry_creditors_debit_total_after
                Sundry_creditors_credit_total_after = "-"

            ledgers = AccountLedger.objects.filter(account_chart__group="Sundry Creditors")
            subledger_entries = []

            for ledger in ledgers:
                subledger_data = {
                    'subledger': ledger.ledger_name,
                    'total_value': ledger.total_value
                }
                subledger_entries.append(subledger_data)
                
            new_sundry_creditors_entry = {'ledger': 'Sundry Creditors', 'group': 'Liability', 'debit': Sundry_creditors_debit_total_after, 'credit': Sundry_creditors_credit_total_after, 'subledgers': subledger_entries}
            new_ledgers_liability.append(new_sundry_creditors_entry)


        # Add the new ledger entries to the respective sections
        for entry in trial_balance:
            for groups in entry['groups']:

                if entry['real_account_type'] == 'Asset' and groups['group'] == 'Sundry Debtors':
                    groups['ledgers'] = new_ledgers_asset
                #     entry['group'] = "Sundry Debtors"
                elif entry['real_account_type'] == 'Liability' and groups['group'] == 'Sundry Creditors':
                    groups['ledgers'] = new_ledgers_liability
                    entry['group'] = "Sundry Creditors"

        context = {
            'trial_balance': trial_balance,
            "total": total,
            "from_date":from_date,
            "to_date":to_date,
            'current_fiscal_year':current_fiscal_year,
            'first_date': first_date,
            'last_date': last_date,
            'trial_active' : Organization.objects.first().show_zero_ledgers

        }

        return render(request, 'accounting/trial_balance.html', context)



class ProfitAndLoss(IsAdminMixin, TemplateView):
    template_name = "accounting/profit_and_loss.html"

    def get_context_data(self, **kwargs):
        from_date_str = self.request.GET.get('fromDate', None)
        to_date_str = self.request.GET.get('toDate', None)
        context = super().get_context_data(**kwargs)
        if from_date_str and to_date_str:
            from_date = parse_datetime(from_date_str)
            to_date_frontend = parse_datetime(to_date_str)
            to_date = to_date_frontend + timedelta(days=1)
            if not from_date or not to_date:
                raise ValidationError('Invalid date format. Dates must be in YYYY-MM-DD HH:MM[:ss[.uuuuuu]][TZ] format.')
        else:
            # Fallback to no date filtering if dates are not provided
            from_date = None
            to_date = None

        if from_date and to_date:
            expenses = CumulativeLedger.objects.filter(
                ~Q(total_value=0), 
                account_chart__account_type="Expense", 
                created_at__range=[from_date, to_date]
            )
            revenues = CumulativeLedger.objects.filter(
                ~Q(total_value=0), 
                account_chart__account_type="Revenue", 
                created_at__range=[from_date, to_date]
            )
            
            liabilities = CumulativeLedger.objects.filter(
                ~Q(total_value=0), 
                account_chart__account_type="Liability", 
                created_at__range=[from_date, to_date]
            )
            assets = CumulativeLedger.objects.filter(
                ~Q(total_value=0), 
                account_chart__account_type="Asset", 
                created_at__range=[from_date, to_date]
            )
        else:
            expenses = CumulativeLedger.objects.filter(
                ~Q(total_value=0), 
                account_chart__account_type="Expense"
            )
            revenues = CumulativeLedger.objects.filter(
                ~Q(total_value=0), 
                account_chart__account_type="Revenue"
            )
            liabilities = CumulativeLedger.objects.filter(
                ~Q(total_value=0), 
                account_chart__account_type="Liability"
            )
            assets = CumulativeLedger.objects.filter(
                ~Q(total_value=0), 
                account_chart__account_type="Asset"
            )

        expense_aggregated = expenses.values('ledger_id','ledger_name').annotate(
            total_value=Sum('value_changed')
        ).order_by('ledger_name')

        revenue_aggregated = revenues.values('ledger_id','ledger_name').annotate(
            total_value=Sum('value_changed')
        ).order_by('ledger_name')
        liability_aggregated = liabilities.values('ledger_id','ledger_name').annotate(
            total_value=Sum('value_changed')
        ).order_by('ledger_name')
        asset_aggregated = assets.values('ledger_id','ledger_name').annotate(
            total_value=Sum('value_changed')
        ).order_by('ledger_name')

        # Convert querysets to lists
        expense_list = list(expense_aggregated)
        revenue_list = list(revenue_aggregated)
        liability_list = list(liability_aggregated)
        asset_list = list(asset_aggregated)

        print(expense_list)
        expense_list, expense_total, revenue_list, revenue_total, liability_list, liability_total, asset_list, asset_total = ProfitAndLossData.get_data(revenues=revenue_list, expenses=expense_list, liabilities=liability_list, assets=asset_list)


        context['expenses'] = expense_list
        context['expense_total'] = expense_total
        context['revenues'] = revenue_list
        context['revenue_total'] = revenue_total
        context['liabilities'] = liability_list
        context['liability_total'] = liability_total
        context['assets'] = asset_list
        context['asset_total'] = asset_total
        
        differenceForProfitorLoss = revenue_total + asset_total - expense_total - liability_total

        context['differenceForProfitorLoss'] = differenceForProfitorLoss

        return context

from .utils import BalanceSheetData
class BalanceSheet(IsAdminMixin, TemplateView):
    template_name = "accounting/balance_sheet.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        asset_dict = {}
        liability_dict = {}

        assets = AccountChart.objects.filter(account_type='Asset')
        for ledger in assets:
            sub = AccountLedger.objects.filter(~Q(total_value=0), account_chart__group=ledger)
            if sub:
                asset_dict[ledger.group] = sub


        liabilities = AccountChart.objects.filter(Q(account_type="Liability") | Q(account_type="Equity") )
        for ledger in liabilities:
            sub = AccountLedger.objects.filter(~Q(total_value=0), account_chart__group=ledger)
            if sub:
                liability_dict[ledger.group] = sub

        asset_total = AccountLedger.objects.filter(account_chart__account_type='Asset').aggregate(Sum('total_value')).get('total_value__sum')
        liability_total = AccountLedger.objects.filter(Q(account_chart__account_type="Liability") | Q(account_chart__account_type="Equity") )\
                                    .aggregate(Sum('total_value')).get('total_value__sum')
        

        """"""
        expenses = AccountLedger.objects.filter(~Q(total_value=0), account_chart__account_type="Expense")
        revenues = AccountLedger.objects.filter(~Q(total_value=0), account_chart__account_type="Revenue")
        _, expense_total, _, revenue_total = BalanceSheetData.get_data(revenues=revenues, expenses=expenses)

        sundry_debtors = AccountLedger.objects.filter(
            ~Q(total_value=0),
            account_chart__group="Sundry Debtors"
        ).aggregate(Sum('total_value'))['total_value__sum']

        # Calculate the total value for Sundry Creditors
        sundry_creditors = AccountLedger.objects.filter(
            ~Q(total_value=0),
            account_chart__group="Sundry Creditors"
        ).aggregate(Sum('total_value'))['total_value__sum']


        # print(revenue_total)
        # print(expense_total)
        """"""
        # print(f"Revenue_total {revenue_total}")
        # print(f"Expense_total {expense_total}")
        if revenue_total > expense_total:
            difference_sum = abs(revenue_total-expense_total)
            context['lib_retained_earnings'] =  difference_sum
            # context['lib_retained_earnings'] =  revenue_total
            liability_total +=  difference_sum  #difference of expense and the sales
            # print(liability_total)
        else:
            difference_sum = abs(revenue_total-expense_total)
            context['asset_retained_earnings'] =  difference_sum
            asset_total +=  difference_sum
            # print(liability_total)
        
        context['asset_total'] = asset_total
        context['liability_total'] = liability_total
        context['assets'] = asset_dict
        context['liabilities'] =  liability_dict
        context['sundry_debtors_total'] = sundry_debtors
        context['sundry_creditors_total'] = sundry_creditors

        return context


class DepreciationView(IsAdminMixin, View):

    def get(self, request):
        depreciations = Depreciation.objects.all()
        return render(request, 'accounting/depreciation_list.html', {'depreciations':depreciations})
    
# class PartyLedgerView(IsAdminMixin, View):
#     template_name = 'accounting/partyledger_list.html'
  
    # def get(self, request):
    #     # depreciations = Depreciation.objects.all()
    #     return render(request, 'accounting/partyledger_list.html')

    # def get(self, request):
    #     depreciations = Depreciation.objects.all()
    #     return render(request, 'accounting/depreciation_list.html', {'depreciations':depreciations})

class PartyLedgerView(View):
    template_name = 'accounting/partyledger_list.html'

    def get(self, request):
        search_query = request.GET.get('ledger_search', '')
        ledgers = AccountLedger.objects.filter(ledger_name__icontains=search_query)
        all_ledger_names = list(AccountLedger.objects.values_list('ledger_name', flat=True))

        context = {
            'ledgers': ledgers,
            'search_query': search_query,
            'ledger_names': all_ledger_names,
        }

        return render(request, self.template_name, context)
    
#for paying payment debit ledgers


from bill.utils import update_cumulative_ledger_partyledger 
class PartyLedgerJournalView(CreateView):
    template_name = 'accounting/partyledgerjournal.html'

    def get(self, request, ledger_id):
     
        paid_from = AccountLedger.objects.filter(account_chart__group='Liquid Asset')
        
        # search_query = request.GET.get('ledger_search', '')
        paying_ledger = AccountLedger.objects.get(id=ledger_id)
      
        

        context = {
            'ledger': paying_ledger,
            'paid_from':paid_from
          
        }

        return render(request, self.template_name, context)
    
    def post(self, request, ledger_id):

        debit_ledger1 = request.POST.get('debit_ledger')
        selected_ledger = AccountLedger.objects.get(id=debit_ledger1)
        debit_sub_ledger1 = request.POST.get('subledger')
        print(f" data {request.POST}")
        # print(selected_ledger)
        debit_ledger1 = request.POST.get('debit_ledger')
        # print(debit_ledger1)
        amount = request.POST.get('amount')
        particular = request.POST.get('particular')
        # journal_entry = TblJournalEntry.objects.create(employee_name=request.user.username, journal_total=amount)
        # print(journal_entry)
        entry_date = datetime.now()
        journal_entry = TblJournalEntry.objects.create(employee_name=request.user.username, journal_total=amount, entry_date=entry_date.date())
       
        debit_ledger_id = ledger_id
        debit_ledger = AccountLedger.objects.get(pk=debit_ledger_id)
        print(debit_ledger_id)

        debit_particular = particular
        debit_amount = Decimal(amount)
        debit_ledger_type = debit_ledger.account_chart.account_type
   

        TblDrJournalEntry.objects.create(ledger=debit_ledger, journal_entry=journal_entry, particulars=debit_particular, debit_amount=debit_amount, paidfrom_ledger=selected_ledger)
        if debit_ledger_type in ['Asset', 'Expense']:
            debit_ledger.total_value =debit_ledger.total_value + debit_amount
            debit_ledger.save()
            # update_cumulative_ledger_bill(debit_ledger, entry_date)
            update_cumulative_ledger_partyledger(debit_ledger, entry_date, journal_entry)
                

        elif debit_ledger_type in ['Liability', 'Revenue', 'Equity']:
            debit_ledger.total_value = debit_ledger.total_value - debit_amount
            debit_ledger.save()
            # update_cumulative_ledger_bill(debit_ledger, entry_date)    
            update_cumulative_ledger_partyledger(debit_ledger, entry_date, journal_entry)

        
        credit_ledger_id = debit_ledger1
        # print(credit_ledger_id)
        # credit_ledger1 = AccountLedger.objects.get(pk=debit_ledger)
        # print(credit_ledger1)
        credit_ledger = AccountLedger.objects.get(id=credit_ledger_id)
        print(credit_ledger)
        credit_particular = particular
        credit_amount = D(amount)
        credit_ledger_type = credit_ledger.account_chart.account_type
        TblCrJournalEntry.objects.create(ledger=credit_ledger, journal_entry=journal_entry, particulars=credit_particular, credit_amount=credit_amount, paidfrom_ledger=selected_ledger)
        if credit_ledger_type in ['Asset', 'Expense']:
            credit_ledger.total_value = credit_ledger.total_value - credit_amount
            credit_ledger.save()
            # update_cumulative_ledger_bill(credit_ledger, entry_date)
            update_cumulative_ledger_partyledger(debit_ledger, entry_date, journal_entry)

        elif credit_ledger_type in ['Liability', 'Revenue', 'Equity']:
            credit_ledger.total_value = credit_ledger.total_value + credit_amount
            credit_ledger.save()
            # update_cumulative_ledger_bill(credit_ledger, entry_date)
            update_cumulative_ledger_partyledger(debit_ledger, entry_date, journal_entry)

        current_page_url = reverse('ledger_detail', args=[ledger_id]) + f'?debit_ledger1={debit_ledger1}'
        
        if debit_sub_ledger1:
            selected_sub_ledger = AccountSubLedger.objects.get(id=debit_sub_ledger1)
            print(f"{selected_sub_ledger}- {selected_sub_ledger.sub_ledger_name}")
            prev_value = selected_sub_ledger.total_value
            credit_subledgertracking = AccountSubLedgerTracking.objects.create(subledger = selected_sub_ledger, prev_amount= prev_value, journal=journal_entry)
            selected_sub_ledger.total_value -= decimal.Decimal(D(amount))
            selected_sub_ledger.save()
            credit_subledgertracking.new_amount=selected_sub_ledger.total_value
            credit_subledgertracking.value_changed = selected_sub_ledger.total_value - prev_value
            credit_subledgertracking.save()
        return redirect(current_page_url)
    
#for receiving payment debit ledgers
class PartyLedgerJournal1View(CreateView):
    template_name = 'accounting/partyledgerjournal.html'

    def get(self, request, ledger_id):
     
        paid_from = AccountLedger.objects.filter(account_chart__group='Liquid Asset')
        
        # search_query = request.GET.get('ledger_search', '')
        paying_ledger = AccountLedger.objects.get(id=ledger_id)
      
        

        context = {
            'ledger': paying_ledger,
            'paid_from':paid_from
          
        }

        return render(request, self.template_name, context)
    
    def post(self, request, ledger_id):
        

        debit_ledger1 = request.POST.get('debit_ledger')
        selected_ledger = AccountLedger.objects.get(id=debit_ledger1)
        # print(selected_ledger)
        debit_sub_ledger1 = request.POST.get('subledger')

        print(f" data {request.POST}")
        
        amount = request.POST.get('amount')
        particular = request.POST.get('particular')
        # journal_entry = TblJournalEntry.objects.create(employee_name=request.user.username, journal_total=amount)
        # print(journal_entry)
        entry_date = datetime.now()
        journal_entry = TblJournalEntry.objects.create(employee_name=request.user.username, journal_total=amount, entry_date=entry_date.date())
       
        # debit_ledger_id = ledger_id
        debit_ledger = AccountLedger.objects.get(pk=debit_ledger1)
     

        debit_particular = particular
        debit_amount = Decimal(amount)
        debit_ledger_type = debit_ledger.account_chart.account_type
   

        TblDrJournalEntry.objects.create(ledger=debit_ledger, journal_entry=journal_entry, particulars=debit_particular, debit_amount=debit_amount, paidfrom_ledger=selected_ledger)
        if debit_ledger_type in ['Asset', 'Expense']:
            debit_ledger.total_value =debit_ledger.total_value + debit_amount
            debit_ledger.save()
            # update_cumulative_ledger_bill(debit_ledger, entry_date) 
            update_cumulative_ledger_partyledger(debit_ledger, entry_date, journal_entry)


        elif debit_ledger_type in ['Liability', 'Revenue', 'Equity']:
            debit_ledger.total_value = debit_ledger.total_value - debit_amount
            debit_ledger.save()
            # update_cumulative_ledger_bill(debit_ledger, entry_date)   
            update_cumulative_ledger_partyledger(debit_ledger, entry_date, journal_entry)
 
        
        credit_ledger_id = ledger_id
        # print(credit_ledger_id)
        # credit_ledger1 = AccountLedger.objects.get(pk=debit_ledger)
        # print(credit_ledger1)
        credit_ledger = AccountLedger.objects.get(id=credit_ledger_id)
        # print(credit_ledger)
        credit_particular = particular
        credit_amount = D(amount)
        credit_ledger_type = credit_ledger.account_chart.account_type
        TblCrJournalEntry.objects.create(ledger=credit_ledger, journal_entry=journal_entry, particulars=credit_particular, credit_amount=credit_amount, paidfrom_ledger=selected_ledger)
        if credit_ledger_type in ['Asset', 'Expense']:
            credit_ledger.total_value = credit_ledger.total_value - credit_amount
            credit_ledger.save()
            # update_cumulative_ledger_bill(credit_ledger, entry_date)
            update_cumulative_ledger_partyledger(debit_ledger, entry_date, journal_entry)

        elif credit_ledger_type in ['Liability', 'Revenue', 'Equity']:
            credit_ledger.total_value = credit_ledger.total_value + credit_amount
            credit_ledger.save()
            # update_cumulative_ledger_bill(credit_ledger, entry_date)
            update_cumulative_ledger_partyledger(debit_ledger, entry_date, journal_entry)

        current_page_url = reverse('ledger_detail', args=[ledger_id]) + f'?debit_ledger1={debit_ledger1}'

        if debit_sub_ledger1:
            selected_sub_ledger = AccountSubLedger.objects.get(id=debit_sub_ledger1)
            prev_value = selected_sub_ledger.total_value
            credit_subledgertracking = AccountSubLedgerTracking.objects.create(subledger = selected_sub_ledger, prev_amount= prev_value, journal=journal_entry)
            selected_sub_ledger.total_value += decimal.Decimal(D(amount))
            selected_sub_ledger.save()
            credit_subledgertracking.new_amount=selected_sub_ledger.total_value
            credit_subledgertracking.value_changed = selected_sub_ledger.total_value - prev_value
            credit_subledgertracking.save()
        return redirect(current_page_url)
    
from datetime import datetime, timedelta
from django.db.models import Sum, F, DecimalField, Q, Value
from django.db.models import Min, Max
from django.utils import timezone  as django_timezone
from pytz import timezone as pytz_timezone


class LedgerDetailView(View):
    template_name = 'accounting/ledger_detail.html'  # Replace with your actual template path

    def get(self, request, ledger_id):
        kathmandu_timezone = pytz_timezone('Asia/Kathmandu')
        ledger = get_object_or_404(AccountLedger, id=ledger_id)
        current_date = datetime.today().date()
        tomorrow_date = current_date + timedelta(days=1) 
        # credit_entries = TblCrJournalEntry.objects.filter(ledger=ledger, created_at__range=[current_date, tomorrow_date])
        credit_entries = TblCrJournalEntry.objects.filter(ledger=ledger)

        # credit_entries = TblCrJournalEntry.objects.filter(ledger=ledger)
        total_credit = 0
        total_debit = 0
        # debit_entries = TblDrJournalEntry.objects.filter(ledger=ledger, created_at__range=[current_date, tomorrow_date])
        debit_entries = TblDrJournalEntry.objects.filter(ledger=ledger)

        # debit_entries = TblDrJournalEntry.objects.filter(ledger=ledger)
        from_date = request.GET.get('fromDate')
        to_date = request.GET.get('toDate')
        print(to_date)
        
        option = request.GET.get('option')
        current_fiscal_year = Organization.objects.last().current_fiscal_year
        if from_date and to_date:
            to_date_str = request.GET.get('toDate')
            
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
            to_date += timedelta(days=1)
            credit_entries = TblCrJournalEntry.objects.filter(ledger=ledger)
            debit_entries = TblDrJournalEntry.objects.filter(ledger=ledger)
            credit_entries = credit_entries.filter(journal_entry__entry_date__range=[from_date, to_date])
            debit_entries = debit_entries.filter(journal_entry__entry_date__range=[from_date, to_date])
        unique_journal_ids = debit_entries.values_list('journal_entry_id', flat=True).distinct()
        unique_journal_ids1 = credit_entries.values_list('journal_entry_id', flat=True).distinct()
      
       
        details = []
        for journal_id1 in unique_journal_ids1:
           
            debit_entries_testing = TblDrJournalEntry.objects.filter(journal_entry_id=journal_id1)
            
                
            ledger_names = [debit_entry.ledger.ledger_name for debit_entry in debit_entries_testing]
            ledger_amount = [debit_entry.debit_amount for debit_entry in debit_entries_testing]
         

            debit_entries_test3 = TblCrJournalEntry.objects.filter(Q(journal_entry_id=journal_id1) & Q(ledger_id=ledger_id))
            # date = [debit_entry.created_at.astimezone(kathmandu_timezone).strftime("%Y-%m-%d %H:%M:%S") for debit_entry in debit_entries_test3]
            date = [debit_entry.journal_entry.entry_date.strftime("%Y-%m-%d") for debit_entry in debit_entries_test3]
            credit  = [debit_entry.credit_amount for debit_entry in debit_entries_test3]
            # print(credit)
            particulars  = [debit_entry.particulars for debit_entry in debit_entries_test3]
           
            payers_info = ', '.join([f'{name} ({amount})' for name, amount in zip(ledger_names, ledger_amount)])
           
           
            for credit_entry in credit_entries:
                
             
                # Create a dictionary for this ledger
                debit_ledger_entry = {
                        'journal_id': journal_id1,
                        # 'payers': ', '.join(ledger_names),  # Join ledger names with a comma
                        'payers': payers_info,
                        'date': date,
                        'particulars': particulars,
                        'debit': Decimal('0'),
                        'credit': credit

                    }

            details.append(debit_ledger_entry)

        for journal_id1 in unique_journal_ids:
            debit_entries_test1 = TblCrJournalEntry.objects.filter(journal_entry_id=journal_id1)
            ledger_names = [debit_entry.ledger.ledger_name for debit_entry in debit_entries_test1]
            ledger_amount = [debit_entry.credit_amount for debit_entry in debit_entries_test1]
            payers_info = ', '.join([f'{name} ({amount})' for name, amount in zip(ledger_names, ledger_amount)])

            debit_entries_test2 = TblDrJournalEntry.objects.filter(Q(journal_entry_id=journal_id1) & Q(ledger_id=ledger_id))
            date = [debit_entry.journal_entry.entry_date.strftime("%Y-%m-%d")  for debit_entry in debit_entries_test2]
            debit  = [debit_entry.debit_amount for debit_entry in debit_entries_test2]
            particulars  = [debit_entry.particulars for debit_entry in debit_entries_test2]

            for debit_entry in debit_entries:
             
                # Create a dictionary for this ledger
                debit_ledger_entry = {
                        'journal_id': journal_id1,
                        # 'payers': ', '.join(ledger_names),  # Join ledger names with a comma
                        'payers': payers_info,
                        'date': date,
                        'particulars': particulars,
                        'debit': debit,
                        'credit': Decimal('0')

                    }

            details.append(debit_ledger_entry)
        
  

        sorted_details = sorted(details, key=lambda x: x['date'])




        # Check the count of credit entries
        credit_entry_count = credit_entries.count()

        # Check the count of debit entries
        debit_entry_count = debit_entries.count()

        # Initialize variables as None
        first_date = None
        last_date = None


        if credit_entry_count > 0 or debit_entry_count > 0:
            credit_date_range = credit_entries.aggregate(
                first_credit_date=Min('journal_entry__entry_date'),
                last_credit_date=Max('journal_entry__entry_date')
            )

            # Calculate the first and last dates for debit entries
            debit_date_range = debit_entries.aggregate(
                first_debit_date=Min('journal_entry__entry_date'),
                last_debit_date=Max('journal_entry__entry_date')
            )

            from datetime import date 
            first_credit_date = credit_date_range['first_credit_date'] or date.max
            last_credit_date = credit_date_range['last_credit_date'] or date.min
            first_debit_date = debit_date_range['first_debit_date'] or date.max
            last_debit_date = debit_date_range['last_debit_date'] or date.min

            # Determine the overall date range
            first_date = min(first_credit_date, first_debit_date) 
            last_date = max(last_credit_date, last_debit_date) 

        print('first_date_from_loop', first_date)
        if from_date and to_date:
            credit_entries = credit_entries.filter(journal_entry__entry_date__range=[from_date, to_date])
            debit_entries = debit_entries.filter(journal_entry__entry_date__range=[from_date, to_date])

        for credit in credit_entries:
            total_credit += credit.credit_amount

        all_entries = sorted(
            chain(credit_entries, debit_entries),
            key=attrgetter('created_at')
        )

        for debit in debit_entries:
            total_debit += debit.debit_amount

        total_diff = total_debit - total_credit
        opening_balance_credit = 0
        opening_balance_debit = 0
        closing_balance = 0

        yesterday = current_date - timedelta(days=1)

        earliest_credit_date = TblCrJournalEntry.objects.aggregate(earliest_credit_date=Min('journal_entry__entry_date'))['earliest_credit_date']
        earliest_debit_date = TblDrJournalEntry.objects.aggregate(earliest_debit_date=Min('journal_entry__entry_date'))['earliest_debit_date']

        
        if earliest_credit_date is not None and earliest_debit_date is not None:
            # Determine the overall first date
            first_date1 = min(earliest_credit_date, earliest_debit_date)

            print('first_date_from_method', first_date1)

            # Filter credit entries from 'first_date' to 'yesterday'
            opening_credit_entries = TblCrJournalEntry.objects.filter(ledger=ledger, created_at__range=[first_date1, current_date])

            # Aggregate the sum of credit amounts
            opening_balance_credit = opening_credit_entries.aggregate(total_credit=Sum('credit_amount'))['total_credit'] or 0

            opening_debit_entries = TblDrJournalEntry.objects.filter(ledger=ledger, created_at__range=[first_date1, current_date])

            # Aggregate the sum of credit amounts
            opening_balance_debit = opening_debit_entries.aggregate(total_debit=Sum('debit_amount'))['total_debit'] or 0
        else:
            opening_balance_credit = None
            opening_balance_debit = None
        # first_date1 = min(earliest_credit_date, earliest_debit_date)

        # print('first_date_from_method', first_date1)

        # # Filter credit entries from 'first_date' to 'yesterday'
        # opening_credit_entries = TblCrJournalEntry.objects.filter(ledger=ledger, created_at__range=[first_date1, current_date])

        # # Aggregate the sum of credit amounts
        # opening_balance_credit = opening_credit_entries.aggregate(total_credit=Sum('credit_amount'))['total_credit'] or 0

        # opening_debit_entries = TblDrJournalEntry.objects.filter(ledger=ledger, created_at__range=[first_date1, current_date])

        # # Aggregate the sum of credit amounts
        # opening_balance_debit = opening_debit_entries.aggregate(total_debit=Sum('debit_amount'))['total_debit'] or 0

        
        if option == 'openclose':
            total_diff1 = total_debit - total_credit
            if total_diff1 < 0:
                opening_balance_credit = abs(total_diff1)
            else:
                closing_balance = total_diff1

        neg = 0
        if total_diff < 0:
            neg = 1
        elif total_diff == 0:
            neg = 2
        closing_balance = abs(total_diff)

        # Calculate the opening balance before filtering
        if from_date:
            to_date_str = request.GET.get('toDate')        
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
            to_date += timedelta(days=1)
            opening_balance_credit = 0
            opening_balance_debit = 0
            opening_entries_credit = TblCrJournalEntry.objects.filter(ledger=ledger, journal_entry__entry_date__lt=from_date)
            opening_entries_debit = TblDrJournalEntry.objects.filter(ledger=ledger, journal_entry__entry_date__lt=from_date)
            
            for entry in opening_entries_credit:
                opening_balance_credit += entry.credit_amount
            for entry in opening_entries_debit:
                opening_balance_debit += entry.debit_amount
            to_date -= timedelta(days=1)

            to_date = to_date.strftime('%Y-%m-%d')

        context = {
            'ledger': ledger,
            'entries': all_entries,
            'credit_entries': credit_entries,
            'debit_entries': debit_entries,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'closing_balance': closing_balance,
            'opening_balance': {'credit': opening_balance_credit if opening_balance_debit else 0, 'debit': opening_balance_debit if opening_balance_debit else 0},
            'from_date': from_date,
            'to_date': to_date,
            'openclose': option == 'openclose',
            'neg': neg,
            'current_fiscal_year':current_fiscal_year,
            'first_credit_date': first_date,
            'last_credit_date': last_date,
            'sorted_details':sorted_details
    
        }
        return render(request, self.template_name, context)
    


@api_view(['POST'])
def end_fiscal_year(request):
        org = Organization.objects.first()
        fiscal_year = org.get_fiscal_year()
        ledgers = AccountLedger.objects.all()
        sub_ledgers = AccountSubLedger.objects.all()
        accumulated_depn = AccountLedger.objects.get(ledger_name='Accumulated Depreciation')


        for sub in sub_ledgers:
            FiscalYearSubLedger.objects.create(sub_ledger_name=sub.sub_ledger_name, total_value=sub.total_value, fiscal_year=fiscal_year, ledger=sub.ledger)

        for led in ledgers:
            FiscalYearLedger.objects.create(ledger_name=led.ledger_name, total_value=led.total_value,fiscal_year=fiscal_year, account_chart=led.account_chart)
            if led.account_chart.account_type in ['Revenue', 'Expense']:
                for sub in led.accountsubledger_set.all():
                    if sub.ledger.account_chart.group == 'Depreciation':
                        accumulated_depn.total_value += sub.total_value
                        accumulated_depn.save()
                    sub.total_value=0
                    sub.save()
                if not led.ledger_name == 'Accumulated Depreciation':
                    led.total_value = 0
                    led.save()
                    # AccountSubLedger.objects.create(sub_ledger_name=f'{sub.sub_ledger_name} for {fiscal_year}', total_value=sub.total_value, ledger=accumulated_depn)
                    
        depreciations = Depreciation.objects.filter(fiscal_year=fiscal_year)

        org.start_year+=1
        org.end_year += 1

        org.save()

        for depn in depreciations:
            amount = float(depn.net_amount)
            percentage = depn.item.asset.depreciation_pool.percentage
            bill_date = depn.item.asset_purchase.bill_date
            depreciation_amount, bs_date = calculate_depreciation(amount, percentage, bill_date)
            net_amount = amount-depreciation_amount
            Depreciation.objects.create(miti=bs_date,depreciation_amount=depreciation_amount, net_amount=net_amount, item=depn.item, ledger=depn.ledger)
            depreciation_amount = D(depreciation_amount)
            depn_subledger = AccountSubLedger.objects.get(sub_ledger_name=f'{depn.item.asset.title} Depreciation')
            depn_subledger.total_value += depreciation_amount
            depn_subledger.save()


            depn_ledger = depn_subledger.ledger
            depn_ledger.total_value+= depreciation_amount
            depn_ledger.save()

            asset_ledger = depn.ledger
            asset_ledger.total_value -= depreciation_amount
            asset_ledger.save()

            asset_ledger = AccountSubLedger.objects.get(sub_ledger_name=depn.item.asset.title, ledger__account_chart__account_type='Asset')
            asset_ledger.total_value -= depreciation_amount
            asset_ledger.save()
        
        return Response({})



from django.db.models import Sum



class SundryDebtorsLedgersView(View):
    template_name = 'accounting/sundry_debtors.html'

    def get(self, request):
        # Query the AccountLedger model to get all ledgers with group "Sundry Debtors"
        sundry_debtors_ledgers = AccountLedger.objects.filter(account_chart__group="Sundry Debtors").order_by('ledger_name')

        # Get the filter parameters from the request
        from_date = request.GET.get('fromDate')
        to_date = request.GET.get('toDate')
        option = request.GET.get('option')
        current_fiscal_year = Organization.objects.last().current_fiscal_year

        # Create a list to store ledger details
        ledger_details = []

        total_d = 0  # for storing the all total of the debit of the entries in ledger_details
        total_c = 0  # for storing the all total of the credit of the entries in ledger_details

        # Initialize variables for overall date range
        first_date = None
        last_date = None

        # Calculate debit and credit totals for each ledger for the selected date
        for ledger in sundry_debtors_ledgers:
            # Filter debit and credit entries by date range
            debit_entries = TblDrJournalEntry.objects.filter(ledger=ledger)
            credit_entries = TblCrJournalEntry.objects.filter(ledger=ledger)

            # Check the count of credit entries
            credit_entry_count = credit_entries.count()

            # Check the count of debit entries
            debit_entry_count = debit_entries.count()

            if credit_entry_count > 0 or debit_entry_count > 0:
                # Calculate the first and last dates for credit entries
                credit_date_range = credit_entries.aggregate(
                    first_credit_date=Min('created_at'),
                    last_credit_date=Max('created_at')
                )

                # Calculate the first and last dates for debit entries
                debit_date_range = debit_entries.aggregate(
                    first_debit_date=Min('created_at'),
                    last_debit_date=Max('created_at')
                )

                # Determine the minimum date for this ledger (comparing debit and credit)
                ledger_first_date = None

                if credit_date_range['first_credit_date'] and debit_date_range['first_debit_date']:
                    ledger_first_date = min(
                        credit_date_range['first_credit_date'],
                        debit_date_range['first_debit_date']
                    )
                elif credit_date_range['first_credit_date']:
                    ledger_first_date = credit_date_range['first_credit_date']
                elif debit_date_range['first_debit_date']:
                    ledger_first_date = debit_date_range['first_debit_date']

                # Update the overall minimum date
                if ledger_first_date and (first_date is None or ledger_first_date < first_date):
                    first_date = ledger_first_date

                # Determine the maximum date for this ledger (comparing debit and credit)
                ledger_last_date = None

                if credit_date_range['last_credit_date'] and debit_date_range['last_debit_date']:
                    ledger_last_date = max(
                        credit_date_range['last_credit_date'],
                        debit_date_range['last_debit_date']
                    )
                elif credit_date_range['last_credit_date']:
                    ledger_last_date = credit_date_range['last_credit_date']
                elif debit_date_range['last_debit_date']:
                    ledger_last_date = debit_date_range['last_debit_date']

                # Update the overall maximum date
                if ledger_last_date and (last_date is None or ledger_last_date > last_date):
                    last_date = ledger_last_date

            # Rest of your code for calculating totals and appending to ledger_details
            if from_date and to_date:
                from_date = request.GET.get('fromDate')
                to_date_str = request.GET.get('toDate')        
                to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
                to_date += timedelta(days=1)
                debit_entries = debit_entries.filter(created_at__range=[from_date, to_date])
                credit_entries = credit_entries.filter(created_at__range=[from_date, to_date])
                to_date -= timedelta(days=1)
                to_date = to_date.strftime('%Y-%m-%d')

            total_credit=0
            total_debit=0
                # Calculate debit total for the selected date
            for credit in credit_entries:
                total_credit += credit.credit_amount

                # Calculate credit total for the selected date
            for debit in debit_entries:
                total_debit += debit.debit_amount

            if total_debit != 0 or total_credit != 0:
                ledger_details.append({
                    'ledger_id': ledger.id,
                    'ledger_name': ledger.ledger_name,
                    'debit_total': total_debit,
                    'credit_total': total_credit,
                })
                    
                    #Update the overall totals
            total_d += total_debit
            total_c += total_credit

        balance = total_d - total_c
        if balance <= 0:
            c = 0  # tells that it is negative
        else:
            c = 1  # tells that it is positive

        new_balance = abs(balance)

        # Render a template with the retrieved ledger details and filter parameters
        return render(request, self.template_name, {
            'ledger_details': ledger_details,
            'from_date': from_date,
            'to_date': to_date,
            'option': option,
            'total_d': total_d,
            'total_c': total_c,
            'openclose': option == 'openclose',
            'balance': new_balance,
            'c': c,
            'current_fiscal_year': current_fiscal_year,
            'first_date': first_date,
            'last_date': last_date
        })

class SundryCreditorsLedgersView(View):
    template_name = 'accounting/sundry_creditors.html'

    def get(self, request):
        # Query the AccountLedger model to get all ledgers with group "Sundry Debtors"
        sundry_creditors_ledgers = AccountLedger.objects.filter(account_chart__group="Sundry Creditors").order_by('ledger_name')

        # Get the filter parameters from the request
        from_date = request.GET.get('fromDate')
        to_date = request.GET.get('toDate')
        option = request.GET.get('option')
        current_fiscal_year = Organization.objects.last().current_fiscal_year

        # Create a list to store ledger details
        ledger_details = []

        total_d = 0  # for storing the all total of the debit of the entries in ledger_details
        total_c = 0  # for storing the all total of the credit of the entries in ledger_details

        # Initialize variables for overall date range
        first_date = None
        last_date = None

        # Calculate debit and credit totals for each ledger for the selected date
        for ledger in sundry_creditors_ledgers:
            # Filter debit and credit entries by date range
            debit_entries = TblDrJournalEntry.objects.filter(ledger=ledger)
            credit_entries = TblCrJournalEntry.objects.filter(ledger=ledger)

            # Check the count of credit entries
            credit_entry_count = credit_entries.count()

            # Check the count of debit entries
            debit_entry_count = debit_entries.count()

            if credit_entry_count > 0 or debit_entry_count > 0:
                # Calculate the first and last dates for credit entries
                credit_date_range = credit_entries.aggregate(
                    first_credit_date=Min('created_at'),
                    last_credit_date=Max('created_at')
                )

                # Calculate the first and last dates for debit entries
                debit_date_range = debit_entries.aggregate(
                    first_debit_date=Min('created_at'),
                    last_debit_date=Max('created_at')
                )

                # Determine the minimum date for this ledger (comparing debit and credit)
                ledger_first_date = None

                if credit_date_range['first_credit_date'] and debit_date_range['first_debit_date']:
                    ledger_first_date = min(
                        credit_date_range['first_credit_date'],
                        debit_date_range['first_debit_date']
                    )
                elif credit_date_range['first_credit_date']:
                    ledger_first_date = credit_date_range['first_credit_date']
                elif debit_date_range['first_debit_date']:
                    ledger_first_date = debit_date_range['first_debit_date']

                # Update the overall minimum date
                if ledger_first_date and (first_date is None or ledger_first_date < first_date):
                    first_date = ledger_first_date

                # Determine the maximum date for this ledger (comparing debit and credit)
                ledger_last_date = None

                if credit_date_range['last_credit_date'] and debit_date_range['last_debit_date']:
                    ledger_last_date = max(
                        credit_date_range['last_credit_date'],
                        debit_date_range['last_debit_date']
                    )
                elif credit_date_range['last_credit_date']:
                    ledger_last_date = credit_date_range['last_credit_date']
                elif debit_date_range['last_debit_date']:
                    ledger_last_date = debit_date_range['last_debit_date']

                # Update the overall maximum date
                if ledger_last_date and (last_date is None or ledger_last_date > last_date):
                    last_date = ledger_last_date

            # Rest of your code for calculating totals and appending to ledger_details
            if from_date and to_date:
                from_date = request.GET.get('fromDate')
                to_date_str = request.GET.get('toDate')        
                to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
                to_date += timedelta(days=1)
                debit_entries = debit_entries.filter(created_at__range=[from_date, to_date])
                credit_entries = credit_entries.filter(created_at__range=[from_date, to_date])
                to_date -= timedelta(days=1)
                to_date = to_date.strftime('%Y-%m-%d')

            total_credit=0
            total_debit=0
                # Calculate debit total for the selected date
            for credit in credit_entries:
                total_credit += credit.credit_amount

                # Calculate credit total for the selected date
            for debit in debit_entries:
                total_debit += debit.debit_amount

            if total_debit != 0 or total_credit != 0:
                ledger_details.append({
                    'ledger_id': ledger.id,
                    'ledger_name': ledger.ledger_name,
                    'debit_total': total_debit,
                    'credit_total': total_credit,
                })
                    
                    #Update the overall totals
            total_d += total_debit
            total_c += total_credit

        balance = total_d - total_c
        if balance <= 0:
            c = 0  # tells that it is negative
        else:
            c = 1  # tells that it is positive

        new_balance = abs(balance)

        # Render a template with the retrieved ledger details and filter parameters
        return render(request, self.template_name, {
            'ledger_details': ledger_details,
            'from_date': from_date,
            'to_date': to_date,
            'option': option,
            'total_d': total_d,
            'total_c': total_c,
            'openclose': option == 'openclose',
            'balance': new_balance,
            'c': c,
            'current_fiscal_year': current_fiscal_year,
            'first_date': first_date,
            'last_date': last_date
        })

from .utils import adjust_cumulative_ledger_afterentries
def soft_delete_journal(request, journal_id):
    try:
        # Retrieve the journal entry or return a 404 if it doesn't exist
        journal_entry = get_object_or_404(TblJournalEntry, id=journal_id)

        # Get related credit and debit entries
        credit_entries = TblCrJournalEntry.objects.filter(journal_entry=journal_entry)
        debit_entries = TblDrJournalEntry.objects.filter(journal_entry=journal_entry)

        # Reverse the ledger operations for credit entries
        for credit_entry in credit_entries:
            ledger = credit_entry.ledger
            ledger_type = ledger.account_chart.account_type

            # Reverse the operation based on ledger type
            if ledger_type in ['Asset', 'Expense']:
                ledger.total_value += credit_entry.credit_amount
            elif ledger_type in ['Liability', 'Revenue', 'Equity']:
                ledger.total_value -= credit_entry.credit_amount

            ledger.save()
            # update_cumulative_ledger_bill(ledger)

        # Reverse the ledger operations for debit entries
        for debit_entry in debit_entries:
            ledger = debit_entry.ledger
            ledger_type = ledger.account_chart.account_type

            # Reverse the operation based on ledger type
            if ledger_type in ['Asset', 'Expense']:
                ledger.total_value -= debit_entry.debit_amount
            elif ledger_type in ['Liability', 'Revenue', 'Equity']:
                ledger.total_value += debit_entry.debit_amount

            ledger.save()
            # update_cumulative_ledger_bill(ledger)
        adjust_cumulative_ledger_afterentries(journal_entry)
 
        journal_entry.delete()







    except TblJournalEntry.DoesNotExist:
        # Handle the case where the journal entry doesn't exist.
        messages.error(request, "Journal Entry not found.")
    except Exception as e:
        # Handle any other exceptions or errors as needed
        messages.error(request, f"An error occurred: {str(e)}")

    return redirect('journal_list')

from .models import AccountSubLedgerTracking
from django.http import Http404
class JournalEntryUpdateView(IsAdminMixin, View):
    # Define your GET method to display the form for updating a journal entry
    def get(self, request, pk):
        journal_entry = get_object_or_404(TblJournalEntry, pk=pk)
        check_cumulative = CumulativeLedger.objects.filter(journal=journal_entry).exists()
        if check_cumulative == False:
                journal_entries = TblJournalEntry.objects.all().order_by('-id')
                messages.error(request, "Automatic journal entry cannot be edited")
                return render(request, 'accounting/journal/journal_list.html', {'journal_entries': journal_entries})
   

        ledgers = AccountLedger.objects.all()
        sub_ledgers = AccountSubLedger.objects.all()

        # Get the related debit and credit entries
        debit_entries = TblDrJournalEntry.objects.filter(journal_entry=journal_entry)
        credit_entries = TblCrJournalEntry.objects.filter(journal_entry=journal_entry)

        # Create a list to store debit and credit entry data
        debit_data = []
        credit_data = []




        # Populate debit_data
        for entry in debit_entries:
            debit_subledgers = []
            for subledger in AccountSubLedger.objects.filter(ledger=entry.ledger):
                debit_subledgers.append({
                    'id' : subledger.id,
                    'subledger_name': subledger.sub_ledger_name
                })
            debit_data.append({
                'id': entry.ledger.id,
                'ledger_name': entry.ledger.ledger_name,
                'debit_amount': entry.debit_amount,
                'particular': entry.particulars,
                'subledger_id': entry.sub_ledger.id if entry.sub_ledger else "------",
                'subledger_name': entry.sub_ledger.sub_ledger_name if entry.sub_ledger else "-----",
                'subledgers': debit_subledgers
            })

        # Populate credit_data
        for entry in credit_entries:
            credit_subledgers = []
            for subledger in AccountSubLedger.objects.filter(ledger=entry.ledger):
                credit_subledgers.append({
                    'id' : subledger.id,
                    'subledger_name': subledger.sub_ledger_name
                })
            credit_data.append({
                'id' : entry.ledger.id,
                'ledger_name': entry.ledger.ledger_name,
                'credit_amount': entry.credit_amount,
                'particular' : entry.particulars,
                'subledger_id': entry.sub_ledger.id if entry.sub_ledger else "--------",
                'subledger_name': entry.sub_ledger.sub_ledger_name if entry.sub_ledger else "-------",
                "subledgers": credit_subledgers
            })

        # print(debit_entries)
        # print(credit_entries)

        # Create an instance of the JournalEntryForm and populate it with values
        form = {
            'debit_ledger': debit_data,
            'credit_ledger': credit_data,
            # 'debit_subledgers': debit_subledgers,
            # 'credit_subledgers': credit_subledgers,
            'journal_total': journal_entry.journal_total,
            'fiscal_year': journal_entry.fiscal_year,
            # 'journal_date': journal_entry.entry_date
        }

        # print(form)

        return render(request, 'accounting/journal/journal_entry_update.html', {
            'ledgers': ledgers,
            'sub_ledgers': sub_ledgers,
            'form': form,
            'journal_entry': journal_entry
        })
               
    def get_subledger(self, subledger, ledger):
        subled = None
        if not subledger.startswith('-'):
            try:
                subledger_id = int(subledger)
                subled = AccountSubLedger.objects.get(pk=subledger_id)
            except ValueError:
                subled = AccountSubLedger.objects.create(sub_ledger_name=subledger, is_editable=True, ledger=ledger)
        return subled

    def post(self, request, pk):
        data = request.POST
        debit_ledgers = data.getlist('debit_ledger', [])
        debit_particulars = data.getlist('debit_particular', [])
        debit_amounts = data.getlist('debit_amount', [])
        debit_subledgers = data.getlist('debit_subledger', [])

        credit_ledgers = data.getlist('credit_ledger', [])
        credit_particulars = data.getlist('credit_particular', [])
        credit_amounts = data.getlist('credit_amount', [])
        credit_subledgers = data.getlist('credit_subledger', [])
        # entry_date = data.get('entry_date')
        # narration = data.get('narration')

        ledgers = AccountLedger.objects.all()
        sub_ledgers = AccountSubLedger.objects.all()

        try:
            parsed_debitamt = [D(i) for i in debit_amounts]
            parsed_creditamt = [D(i) for i in credit_amounts]
        except Exception:
            messages.error(request, "Please Enter a valid amount")
            return render(request, 'accounting/journal/journal_entry_update.html', {'ledgers': ledgers, 'sub_ledgers': sub_ledgers})

        debit_sum, credit_sum = sum(parsed_debitamt), sum(parsed_creditamt)
        if debit_sum != credit_sum:
            messages.error(request, "Debit Total and Credit Total must be equal")
            return render(request, 'accounting/journal/journal_entry_update.html', {'ledgers': ledgers, 'sub_ledgers': sub_ledgers})


        for dr in debit_ledgers:
            if dr.startswith('-'):
                messages.error(request, "Ledger must be selected")
                return render(request, 'accounting/journal/journal_entry_update.html', {'ledgers': ledgers, 'sub_ledgers': sub_ledgers})

        # Retrieve the journal_entry object or raise a 404 error if not found
        try:
            journal_entry = TblJournalEntry.objects.get(pk=pk)
        except TblJournalEntry.DoesNotExist:
            raise Http404("Journal Entry does not exist")

        update_cumulative_ledger_journal(journal_entry, data)


        #Update the value of ledger of old entries
        existing_credit_entriess = TblCrJournalEntry.objects.filter(journal_entry=journal_entry)
        print("credit_entries", existing_credit_entriess)
        for existing_credit_entries in existing_credit_entriess:
            old_credit_ledger = existing_credit_entries.ledger if existing_credit_entries else None
            old_credit_subledger = existing_credit_entries.sub_ledger if existing_credit_entries else None
            print("credit_name", old_credit_ledger.ledger_name)
            old_credit_ledger_type = old_credit_ledger.account_chart.account_type
            old_credit_amount = existing_credit_entries.credit_amount
            if old_credit_ledger_type in ['Asset', 'Expense']:
                print("to be increased", old_credit_amount)
                old_credit_ledger.total_value += old_credit_amount
                old_credit_ledger.save()
                if old_credit_subledger:
                    old_credit_subledger.total_value += old_credit_amount
                    old_credit_subledger.save()
                    # subledgertracking = AccountSubLedgerTracking.objects.filter(subledger=old_credit_subledger, journal=journal_entry)

            elif old_credit_ledger_type in ['Liability', 'Revenue', 'Equity']:
                print("to be decreased", old_credit_amount)
                old_credit_ledger.total_value -= old_credit_amount
                old_credit_ledger.save()
                if old_credit_subledger:
                    old_credit_subledger.total_value -= old_credit_amount
                    old_credit_subledger.save()

        # Update or create credit entries

        credit_to_debit_mapping = {}
        for i in range(len(credit_ledgers)):
            credit_ledger_id = int(credit_ledgers[i])
            credit_ledger = AccountLedger.objects.get(pk=credit_ledger_id)
            credit_particular = credit_particulars[i]
            credit_amount = parsed_creditamt[i]
            subledger = self.get_subledger(credit_subledgers[i], credit_ledger)  # Implement your subledger utility function
            credit_ledger_type = credit_ledger.account_chart.account_type
            
            # Check if there is an existing entry for this ledger, if so, update it, otherwise, create a new one
            existing_entry = existing_credit_entriess.filter(ledger=credit_ledger).first()

            if existing_entry:  
                existing_entry.particulars = credit_particular
                existing_entry.credit_amount = credit_amount
                existing_entry.sub_ledger = subledger
                existing_entry.paidfrom_ledger = credit_ledger
                existing_entry.save()
            else:
                TblCrJournalEntry.objects.create(
                    ledger=credit_ledger,
                    journal_entry=journal_entry,
                    particulars=credit_particular,
                    credit_amount=credit_amount,
                    sub_ledger=subledger,
                    paidfrom_ledger=credit_ledger
                )

            if credit_ledger_type in ['Asset', 'Expense']:
                credit_ledger.total_value -= credit_amount
                credit_ledger.save()
                if subledger:
                    subledger.total_value -= credit_amount
                    subledger.save()
            elif credit_ledger_type in ['Liability', 'Revenue', 'Equity']:
                credit_ledger.total_value += credit_amount
                credit_ledger.save()
                if subledger:

                    subledger.total_value += credit_amount
                    subledger.save()



        #Update the value of ledger of old entries
        existing_debit_entriess = TblDrJournalEntry.objects.filter(journal_entry=journal_entry)
        print("debit_entries", existing_debit_entriess)
        for existing_debit_entries in existing_debit_entriess:
            old_debit_ledger = existing_debit_entries.ledger if existing_debit_entries else None
            old_debit_subledger = existing_debit_entries.sub_ledger if existing_debit_entries else None
            print("debit_name", old_debit_ledger.ledger_name)
            old_debit_ledger_type = old_debit_ledger.account_chart.account_type
            old_debit_amount = existing_debit_entries.debit_amount
            if old_debit_ledger_type in ['Asset', 'Expense']:
                print("to be increased", old_debit_amount)
                old_debit_ledger.total_value -= old_debit_amount
                old_debit_ledger.save()
                if old_debit_subledger:
                    old_debit_subledger.total_value -= old_debit_amount
                    old_debit_subledger.save()
            elif old_debit_ledger_type in ['Liability', 'Revenue', 'Equity']:
                print("to be decreased", old_debit_amount)
                old_debit_ledger.total_value += old_debit_amount
                old_debit_ledger.save()
                if old_debit_subledger:
                    old_debit_subledger.total_value += old_debit_amount
                    old_debit_subledger.save()

        # Delete any existing entries that are no longer present in the form
        existing_credit_entriess.exclude(ledger__id__in=credit_ledgers).delete()

        # Update or create debit entries
        # existing_debit_entries = TblDrJournalEntry.objects.filter(journal_entry=journal_entry)
        for i in range(len(debit_ledgers)):
            debit_ledger_id = int(debit_ledgers[i])
            debit_ledger = AccountLedger.objects.get(pk=debit_ledger_id)
            debit_particular = debit_particulars[i]
            debit_amount = parsed_debitamt[i]
            subledger = self.get_subledger(debit_subledgers[i], debit_ledger)  # Implement your subledger utility function
            debit_ledger_type = debit_ledger.account_chart.account_type
            
            # Check if there is an existing entry for this ledger, if so, update it, otherwise, create a new one
            existing_entry = existing_debit_entriess.filter(ledger=debit_ledger).first()
            if existing_entry:
                existing_entry.particulars = debit_particular
                existing_entry.debit_amount = debit_amount
                existing_entry.sub_ledger = subledger
                existing_entry.paidfrom_ledger = credit_to_debit_mapping.get(credit_ledger)
                existing_entry.save()
            else:
                TblDrJournalEntry.objects.create(
                    ledger=debit_ledger,
                    journal_entry=journal_entry,
                    particulars=debit_particular,
                    debit_amount=debit_amount,
                    sub_ledger=subledger,
                    paidfrom_ledger=credit_to_debit_mapping.get(credit_ledger)
                )

            if debit_ledger_type in ['Asset', 'Expense']:
                debit_ledger.total_value += debit_amount
                debit_ledger.save()
                if subledger:
                    subledger.total_value += debit_amount
                    subledger.save()
            elif debit_ledger_type in ['Liability', 'Revenue', 'Equity']:
                debit_ledger.total_value -= debit_amount
                debit_ledger.save()
                if subledger:
                    subledger.total_value -= debit_amount
                    subledger.save()


        existing_debit_entriess.exclude(ledger__id__in=debit_ledgers).delete()

        # Update journal entry data
        journal_entry.employee_name = request.user.username
        journal_entry.journal_total = debit_sum
        # journal_entry.entry_date = entry_date
        # journal_entry.narration = narration
        journal_entry.save()
        return redirect('journal_list')

from django.shortcuts import redirect
from django.views.generic import View
from django.urls import reverse
from .forms import OpeningCountForm
from decimal import Decimal

class EditOpeningCountView(View):
    template_name = 'accounting/opening_count.html'
    success_url = reverse_lazy('accountchart_list')

    def get(self, request, *args, **kwargs):
        ledger_id = kwargs.get('ledger_id')

        ledger = AccountLedger.objects.get(id=ledger_id)
        ledger_name = ledger.ledger_name

        # opening_count_obj = CumulativeLedger.objects.filter(ledger=ledger).exclude(total_value=0).order_by('created_at').first()
        # opening_count = opening_count_obj.total_value

        opening_count = ledger.opening_count


        form = OpeningCountForm()
        context = {
            "ledger_name": ledger_name, 
            "opening_count": opening_count,
            "form": form
        }

        
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        ledger_id = kwargs.get('ledger_id')
        ledger = AccountLedger.objects.get(id=ledger_id)
        ledger_name = ledger.ledger_name
        form = OpeningCountForm(request.POST)
        if form.is_valid():
            opening_count_form = form.cleaned_data['opening_count']
            print(f"This is the opening_count_form {opening_count_form} ")
            
            print(opening_count_form)
            opening_count_obj = CumulativeLedger.objects.filter(ledger=ledger).exclude(total_value=0).order_by('created_at').first()
                # opening_count_database = opening_count_obj.total_value
            opening_count_database = ledger.opening_count
            if opening_count_obj:
                created_at = opening_count_obj.created_at


                print(f"opening_count from database {opening_count_database}")

                difference = Decimal(str(opening_count_form)) - opening_count_database
                print(f"Difference = {difference}")
                opening_count_obj.value_changed += difference
                if opening_count_obj.debit_amount != 0:
                    opening_count_obj.debit_amount += difference
                if opening_count_obj.credit_amount != 0:
                    opening_count_obj.credit_amount += difference
                opening_count_obj.save()
                ledger_entries = CumulativeLedger.objects.filter(created_at__range = (created_at, datetime.now()))
                for entry in ledger_entries:
                    entry.total_value += difference
                    entry.save()

                ledger.total_value += difference
                ledger.opening_count = Decimal(opening_count_form)
                ledger.save()


                return redirect(self.success_url)
            else:
                ledger.opening_count = Decimal(opening_count_form)
                ledger.total_value = Decimal(opening_count_form)
                ledger.save()
                return redirect(self.success_url)
                
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.views import View
# from .utils import give_detail, get_standard_trial_balance  # Ensure these are correctly imported

class ExportTrialBalance(View):
    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('fromDate', None)
        to_date = request.GET.get('toDate', None)
        option = request.GET.get('option', None)

        print(f"from_date_from_export {from_date}")
        print(f"to_date_from_export {from_date}")
        print(f"option_from_export {option}")

        if from_date and to_date:
            if option and option == 'openclose':
                trial_balance, total = give_detail(from_date, to_date)
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = 'Trial Balance'

                # Define column headers
                headers = ['Ledger', 'Opening', 'Debit', 'Credit', 'Closing']
                for col_num, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # Populate the data
                row_num = 2
                total_opening = 0
                total_debit = 0
                total_credit = 0
                total_closing = 0
                
                for data in trial_balance:
                    real_account_type = data['real_account_type']
                    # Add real_account_type as header
                    sheet.cell(row=row_num, column=1, value=real_account_type)
                    sheet.cell(row=row_num, column=1).font = Font(bold=True)
                    row_num += 1

                    for group in data['groups']:
                        group_name = group['group']
                        # Add group as header
                        sheet.cell(row=row_num, column=1, value=group_name)
                        sheet.cell(row=row_num, column=1).font = Font(bold=True, italic=True)
                        row_num += 1

                        for ledger in group['ledgers']:
                            ledger_name = ledger['ledger']
                            opening = ledger.get('opening', 0)
                            debit = ledger.get('debit', 0)
                            credit = ledger.get('credit', 0)
                            closing = ledger.get('closing', 0)

                            # Handle formatting for debit and credit
                            if debit == '-':
                                debit = 0
                            if credit == '-':
                                credit = 0
                            # Write ledger details
                            sheet.cell(row=row_num, column=1, value=ledger_name)
                            sheet.cell(row=row_num, column=2, value=opening)
                            sheet.cell(row=row_num, column=3, value=debit)
                            sheet.cell(row=row_num, column=4, value=credit)
                            sheet.cell(row=row_num, column=5, value=closing)

                            total_opening += opening
                            total_debit += debit
                            total_credit += credit
                            total_closing += closing

                            row_num += 1

                            # Add subledgers if available
                            for subledger in ledger.get('subledgers', []):
                                subledger_name = subledger['subledger']
                                subledger_value = subledger['total_value']

                                sheet.cell(row=row_num, column=1, value=f"  {subledger_name}")
                                sheet.cell(row=row_num, column=2, value=subledger_value)
                                sheet.cell(row=row_num, column=3, value='')
                                sheet.cell(row=row_num, column=4, value='')
                                sheet.cell(row=row_num, column=5, value='')

                                row_num += 1

                        # Add an empty row between groups
                        row_num += 1

                # Add totals
                sheet.cell(row=row_num, column=1, value="Total")
                sheet.cell(row=row_num, column=2, value=total_opening)
                sheet.cell(row=row_num, column=3, value=total_debit)
                sheet.cell(row=row_num, column=4, value=total_credit)
                sheet.cell(row=row_num, column=5, value=total_closing)
                sheet.cell(row=row_num, column=1).font = Font(bold=True)

                # Save the workbook to a BytesIO object
                buffer = io.BytesIO()
                workbook.save(buffer)
                buffer.seek(0)  # Move the cursor to the start of the buffer

                # Create an HTTP response with the Excel file
                response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=Trial_Balance.xlsx'

                return response
            else:
                trial_balance, total = get_standard_trial_balance()
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = 'Trial Balance'

                # Define column headers
                headers = ['Ledger', 'Debit', 'Credit']
                for col_num, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # Populate the data
                row_num = 2
                total_debit = 0
                total_credit = 0
                
                for data in trial_balance:
                    real_account_type = data['real_account_type']
                    # Add real_account_type as header
                    sheet.cell(row=row_num, column=1, value=real_account_type)
                    sheet.cell(row=row_num, column=1).font = Font(bold=True)
                    row_num += 1

                    for group in data['groups']:
                        group_name = group['group']
                        # Add group as header
                        sheet.cell(row=row_num, column=1, value=group_name)
                        sheet.cell(row=row_num, column=1).font = Font(bold=True, italic=True)
                        row_num += 1

                        for ledger in group['ledgers']:
                            ledger_name = ledger['ledger']
                            debit = ledger.get('debit', '-')
                            credit = ledger.get('credit', '-')

                            # Handle formatting for debit and credit
                            if debit == '-':
                                debit = 0
                            if credit == '-':
                                credit = 0
                            
                            # Write ledger details
                            sheet.cell(row=row_num, column=1, value=ledger_name)
                            sheet.cell(row=row_num, column=2, value=debit)
                            sheet.cell(row=row_num, column=3, value=credit)

                            total_debit += debit
                            total_credit += credit

                            row_num += 1

                            # Add subledgers if available
                            for subledger in ledger.get('subledgers', []):
                                subledger_name = subledger['subledger']
                                subledger_value = subledger['total_value']

                                sheet.cell(row=row_num, column=1, value=f"  {subledger_name}")
                                sheet.cell(row=row_num, column=2, value=subledger_value)
                                sheet.cell(row=row_num, column=3, value='')

                                row_num += 1

                        # Add an empty row between groups
                        row_num += 1

                # Add totals
                sheet.cell(row=row_num, column=1, value="Total")
                sheet.cell(row=row_num, column=2, value=total_debit)
                sheet.cell(row=row_num, column=3, value=total_credit)
                sheet.cell(row=row_num, column=1).font = Font(bold=True)

                # Save the workbook to a BytesIO object
                buffer = io.BytesIO()
                workbook.save(buffer)
                buffer.seek(0)  # Move the cursor to the start of the buffer

                # Create an HTTP response with the Excel file
                response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=Trial_Balance.xlsx'

                return response

        else:
            trial_balance, total = get_standard_trial_balance()

            # Create a workbook and set the active worksheet
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Trial Balance'

            # Define column headers
            headers = ['Ledger', 'Debit', 'Credit']
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Populate the data
            row_num = 2
            total_debit = 0
            total_credit = 0
            
            for data in trial_balance:
                real_account_type = data['real_account_type']
                # Add real_account_type as header
                sheet.cell(row=row_num, column=1, value=real_account_type)
                sheet.cell(row=row_num, column=1).font = Font(bold=True)
                row_num += 1

                for group in data['groups']:
                    group_name = group['group']
                    # Add group as header
                    sheet.cell(row=row_num, column=1, value=group_name)
                    sheet.cell(row=row_num, column=1).font = Font(bold=True, italic=True)
                    row_num += 1

                    for ledger in group['ledgers']:
                        ledger_name = ledger['ledger']
                        debit = ledger.get('debit', '-')
                        credit = ledger.get('credit', '-')

                        # Handle formatting for debit and credit
                        if debit == '-':
                            debit = 0
                        if credit == '-':
                            credit = 0
                        
                        # Write ledger details
                        sheet.cell(row=row_num, column=1, value=ledger_name)
                        sheet.cell(row=row_num, column=2, value=debit)
                        sheet.cell(row=row_num, column=3, value=credit)

                        total_debit += debit
                        total_credit += credit

                        row_num += 1

                        # Add subledgers if available
                        for subledger in ledger.get('subledgers', []):
                            subledger_name = subledger['subledger']
                            subledger_value = subledger['total_value']

                            sheet.cell(row=row_num, column=1, value=f"  {subledger_name}")
                            sheet.cell(row=row_num, column=2, value=subledger_value)
                            sheet.cell(row=row_num, column=3, value='')

                            row_num += 1

                    # Add an empty row between groups
                    row_num += 1

            # Add totals
            sheet.cell(row=row_num, column=1, value="Total")
            sheet.cell(row=row_num, column=2, value=total_debit)
            sheet.cell(row=row_num, column=3, value=total_credit)
            sheet.cell(row=row_num, column=1).font = Font(bold=True)

            # Save the workbook to a BytesIO object
            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)  # Move the cursor to the start of the buffer

            # Create an HTTP response with the Excel file
            response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Trial_Balance.xlsx'

            return response
            
import openpyxl
from django.http import HttpResponse
from openpyxl.writer.excel import save_virtual_workbook
from django.db.models import Sum, Q
from .detail_utils import get_subledger_data
from django.utils.dateparse import parse_datetime
from django.core.exceptions import ValidationError
from openpyxl.styles import Alignment
def export_profit_and_loss_to_excel(request):
    from_date_str = request.GET.get('fromDate', None)
    to_date_str = request.GET.get('toDate', None)
    if from_date_str and to_date_str:
        from_date = parse_datetime(from_date_str)
        to_date_frontend = parse_datetime(to_date_str)
        to_date = to_date_frontend + timedelta(days=1)
        if not from_date or not to_date:
            raise ValidationError('Invalid date format. Dates must be in YYYY-MM-DD HH:MM[:ss[.uuuuuu]][TZ] format.')
    else:
        # Fallback to no date filtering if dates are not provided
        from_date = None
        to_date = None

    if from_date and to_date:
        expenses = CumulativeLedger.objects.filter(
            ~Q(total_value=0), 
            account_chart__account_type="Expense", 
            created_at__range=[from_date, to_date]
        )
        revenues = CumulativeLedger.objects.filter(
            ~Q(total_value=0), 
            account_chart__account_type="Revenue", 
            created_at__range=[from_date, to_date]
        )
    else:
        expenses = CumulativeLedger.objects.filter(
            ~Q(total_value=0), 
            account_chart__account_type="Expense"
        )
        revenues = CumulativeLedger.objects.filter(
            ~Q(total_value=0), 
            account_chart__account_type="Revenue"
        )

    # Aggregating results
    expense_aggregated = expenses.values('ledger_name').annotate(
        total_value=Sum('value_changed')
    ).order_by('ledger_name')

    revenue_aggregated = revenues.values('ledger_name').annotate(
        total_value=Sum('value_changed')
    ).order_by('ledger_name')


    # # Create a workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Profit and Loss Statement'

    # Construct header text with date range
    header_text = 'Dhamma Design\n'
    if from_date and to_date:
        header_text += f'(From Date: {from_date.strftime("%Y-%m-%d")} | To Date: {to_date.strftime("%Y-%m-%d")})'
    

    sheet.merge_cells('A1:E1')
    header_cell = sheet.cell(row=1, column=1)
    header_cell.value = header_text
    header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Merge cells for date range
    # date_range_cell = sheet.cell(row=1, column=6)
    # date_range_cell.value = f'From Date: {from_date}\nTo Date: {to_date}' if from_date and to_date else 'Date Range not provided'
    # date_range_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths for better readability
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 30


    sheet.append([])  # Empty row for spacing

    # Write table headers
    sheet.append(['Type', 'Ledger Name', 'Subledger Name', 'Amount'])

    # Function to add ledger and subledger rows to the sheet
    def add_ledger_and_subledger_rows(ledger_aggregated, ledger_type):
        for ledger in ledger_aggregated:
            ledger_name = ledger['ledger_name']
            print(ledger_name)
            total_ledger_value = ledger['total_value']
            
            # Append ledger row
            sheet.append([ledger_type, ledger_name, '', total_ledger_value])
            
            # Fetch subledgers


            subledgers = AccountSubLedger.objects.filter(ledger__ledger_name=ledger_name)

            if from_date and to_date:
                subledger_entries = get_subledger_data(from_date, to_date, subledgers)
            else:
                subledger_entries = []
                for subledger in subledgers:
                    subledger_data = {
                        'subledger': subledger.sub_ledger_name,
                        'total_value': subledger.total_value
                    }
                    subledger_entries.append(subledger_data)
            
            # Append subledger rows
            for subledger in subledger_entries:
                sheet.append(['', '', subledger['subledger'], subledger['total_value']])
    
    # Write revenue data
    add_ledger_and_subledger_rows(revenue_aggregated, 'Revenue')

    # Write expense data
    add_ledger_and_subledger_rows(expense_aggregated, 'Expense')

    # Write totals
    total_revenue = sum(item['total_value'] for item in revenue_aggregated)
    total_expense = sum(item['total_value'] for item in expense_aggregated)
    sheet.append(['Total Revenue', '', '', total_revenue])
    sheet.append(['Total Expense', '', '', total_expense])

    # Create HTTP response with the Excel file
    response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="profit_and_loss_statement.xlsx"'
    return response